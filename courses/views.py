import io
import sys
import openpyxl
from accounts.models import Child, Profile
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.http import HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse, Http404
from .forms import CurriculumProgramForm, CurriculumSyllabusExcelForm, CurriculumSyllabusForm, InstitutionReservationForm, LearningProgramForm, ProgramApplicationForm,ProgramForm, ProgramProductForm, CategoryForm, ProductMaterialFormSet
from .models import Chapter, CurriculumProgram, CurriculumSyllabus, InstitutionReservation, Item, LearningEnrollment, LearningProgram, Program, ProgramApplication, ProgramType, ProgramProduct, Category, ProgramClass, UserProgress
from .forms import ProgramForm, ProgramClassFormSetCreate, ProgramClassFormSetEdit
from .utils import is_child_in_target 
from .utils import safe_exec
from .forms import SyllabusUploadForm
from .models import Program, ProgramSyllabus
from .utils.syllabus_excel import import_syllabus_from_excel
from django.contrib.admin.views.decorators import staff_member_required



@login_required
@user_passes_test(lambda u: u.is_staff)
def program_create(request):
    if request.method == "POST":
        form = ProgramForm(request.POST, request.FILES)
        formset = ProgramClassFormSetCreate(request.POST, queryset=ProgramClass.objects.none())
        if form.is_valid() and formset.is_valid():
            program = form.save()
            classes = formset.save(commit=False)

            for idx, cls in enumerate(classes, start=1):
                cls.program = program
                cls.order = idx

                # ✅ days 문자열 → 리스트 변환
                if isinstance(cls.days, str):
                    cls.days = [d for d in cls.days.split(",") if d]
                elif cls.days is None:
                    cls.days = []

                # ✅ 반 이름 자동 생성
                days_display = ",".join([dict(Program.DAYS_OF_WEEK).get(d, d) for d in cls.days]) if cls.days else ""
                start_hour = cls.start_time.strftime("%H시%M분") if cls.start_time else ""
                cls.name = f"{chr(64+idx)}반-{days_display}{start_hour}"

                cls.save()

            # 삭제 처리
            for deleted in formset.deleted_objects:
                deleted.delete()

            return redirect("program_list")
        else:
            print("폼 에러:", form.errors)
            print("폼셋 에러:", formset.errors)
    else:
        form = ProgramForm()
        # ✅ 새 등록 시에는 기본 1개 행 추가
        formset = ProgramClassFormSetCreate()

    return render(request, "courses/program_form.html", {"form": form, "formset": formset})


from django.contrib import messages

@login_required
@user_passes_test(lambda u: u.is_staff)
def program_edit(request, pk):
    program = get_object_or_404(Program, pk=pk)
    if request.method == "POST":
        form = ProgramForm(request.POST, request.FILES, instance=program)
        formset = ProgramClassFormSetEdit(request.POST, instance=program)

        if form.is_valid() and formset.is_valid():
            program = form.save()
            classes = formset.save(commit=False)

            for idx, cls in enumerate(classes, start=1):
                cls.program = program
                cls.order = idx

                # ✅ DB에는 "mon,wed" 그대로 저장
                # ✅ 반 이름 만들 때만 split
                raw_days = cls.days if isinstance(cls.days, list) else []
                days_display = ",".join([dict(Program.DAYS_OF_WEEK).get(d, d) for d in raw_days])

                start_hour = cls.start_time.strftime("%H시%M분") if cls.start_time else ""
                cls.name = f"{chr(64+idx)}반-{days_display}{start_hour}"
                cls.save()

            # 삭제된 반 처리
            for deleted in formset.deleted_objects:
                deleted.delete()

            messages.success(request, "프로그램이 성공적으로 수정되었습니다 ✅")
            return redirect("program_list")
        else:
            messages.error(request, "저장 중 오류가 발생했습니다. 입력 내용을 확인해주세요 ⚠️")
    else:
        form = ProgramForm(instance=program)
        formset = ProgramClassFormSetEdit(instance=program)

    return render(request, "courses/program_form.html", {
        "form": form,
        "formset": formset,
        "program": program,
    })



from django.db.models import Q, Case, When, Value, IntegerField
from .models import Program, Target, ProgramType

def program_list_always(request):
    programs = Program.objects.filter(recruit_type="always").order_by("-id")
    return render(request, "program/program_list.html", {
        "programs": programs,
        "title": "상시모집 프로그램",
        "active_tab": "always",
    })

def program_list_event(request):
    programs = Program.objects.filter(recruit_type="event").order_by("-id")
    return render(request, "program/program_list.html", {
        "programs": programs,
        "title": "이벤트 프로그램",
        "active_tab": "event",
    })

def program_list_short(request):
    programs = Program.objects.filter(recruit_type="short").order_by("-id")
    return render(request, "program/program_list.html", {
        "programs": programs,
        "title": "단기수업 프로그램",
        "active_tab": "short",
    })

def program_list(request):
    """
    모집 페이지: 카드/표 보기 전환, 검색/필터
    """
    view_type = request.GET.get("view", "cards")  # cards | table
    q = request.GET.get("q", "")
    target = request.GET.get("target", "")
    program_type = request.GET.get("program_type")

    # ForeignKey 최적화
    if request.user.is_authenticated and request.user.is_staff:
        programs = Program.objects.all().select_related("target_start", "target_end")
    elif request.user.is_authenticated and getattr(request.user.profile, "user_type", "") == "center_teacher":
        programs = Program.objects.filter(teacher=request.user.profile) \
            .select_related("target_start", "target_end", "teacher")
    else:
        programs = Program.objects.exclude(status="hidden").select_related("target_start", "target_end")

    # 검색 (프로그램명 / 설명)
    if q:
        programs = programs.filter(Q(name__icontains=q) | Q(description__icontains=q))

    # 대상 필터
    if target:
        programs = programs.filter(
            target_start__id__lte=target,
            target_end__id__gte=target
        )

    # 프로그램 유형 필터
    selected_program_type = None
    if program_type:
        try:
            selected_program_type = int(program_type)
            programs = programs.filter(program_types__id=selected_program_type)
        except ValueError:
            selected_program_type = None

    # ✅ 상태 우선순위 정렬
    status_order = Case(
        When(status="open", then=Value(1)),    # 모집중
        When(status="closed", then=Value(2)),  # 모집마감
        When(status="hidden", then=Value(3)),  # 비공개
        default=Value(4),
        output_field=IntegerField(),
    )

    programs = programs.annotate(status_priority=status_order).order_by("status_priority", "name")

    # 신청 가능 여부
    can_apply = request.user.is_authenticated and (
        request.user.is_superuser or getattr(request.user.profile, "user_type", "") in ["parent", "student"]
    )

    # 드롭다운용 대상 목록
    targets = Target.objects.all().order_by("id")

    context = {
        "programs": programs,
        "view_type": view_type,
        "program_types": ProgramType.objects.all(),
        "q": q,
        "target": target,
        "can_apply": can_apply,
        "targets": targets,
        "program_type": selected_program_type,
    }
    return render(request, "courses/program_list.html", context)

def _get_base_program_queryset(request):
    """기존 program_list 로직 그대로 재사용하는 공통 함수"""
    view_type = request.GET.get("view", "cards")
    q = request.GET.get("q", "")
    target = request.GET.get("target", "")
    program_type = request.GET.get("program_type")

    # 기본 queryset
    if request.user.is_authenticated and request.user.is_staff:
        programs = Program.objects.all().select_related("target_start", "target_end")
    elif request.user.is_authenticated and getattr(request.user.profile, "user_type", "") == "center_teacher":
        programs = Program.objects.filter(teacher=request.user.profile).select_related("target_start", "target_end", "teacher")
    else:
        programs = Program.objects.exclude(status="hidden").select_related("target_start", "target_end")

    # 검색
    if q:
        programs = programs.filter(Q(name__icontains=q) | Q(description__icontains=q))

    # 대상 필터
    if target:
        programs = programs.filter(
            target_start__id__lte=target,
            target_end__id__gte=target
        )

    # 프로그램 유형 필터
    if program_type:
        try:
            program_type_int = int(program_type)
            programs = programs.filter(program_types__id=program_type_int)
        except ValueError:
            pass

    # 모집상태 우선순위 정렬
    status_order = Case(
        When(status="open", then=Value(1)),
        When(status="closed", then=Value(2)),
        When(status="hidden", then=Value(3)),
        default=Value(4),
        output_field=IntegerField(),
    )
    programs = programs.annotate(status_priority=status_order).order_by("status_priority", "name")

    return programs, view_type, q, target, program_type

def program_list_always(request):
    programs, view_type, q, target, program_type = _get_base_program_queryset(request)
    programs = programs.filter(recruit_type="always")

    context = {
        "programs": programs,
        "view_type": view_type,
        "q": q,
        "target": target,
        "program_type": program_type,
        "program_types": ProgramType.objects.all(),
        "page_title": "상시모집 프로그램",
        "active_tab": "always",
    }
    return render(request, "courses/program_list.html", context)


def program_list_event(request):
    programs, view_type, q, target, program_type = _get_base_program_queryset(request)
    programs = programs.filter(recruit_type="event")

    context = {
        "programs": programs,
        "view_type": view_type,
        "q": q,
        "target": target,
        "program_type": program_type,
        "program_types": ProgramType.objects.all(),
        "page_title": "이벤트 프로그램",
        "active_tab": "event",
    }
    return render(request, "courses/program_list.html", context)


def program_list_short(request):
    programs, view_type, q, target, program_type = _get_base_program_queryset(request)
    programs = programs.filter(recruit_type="short")

    context = {
        "programs": programs,
        "view_type": view_type,
        "q": q,
        "target": target,
        "program_type": program_type,
        "program_types": ProgramType.objects.all(),
        "page_title": "단기수업 프로그램",
        "active_tab": "short",
    }
    return render(request, "courses/program_list.html", context)

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from .models import Program, ProgramEnrollment

def program_detail(request, pk):
    program = get_object_or_404(Program, pk=pk)

    profile = getattr(request.user, "profile", None) if request.user.is_authenticated else None
    user_type = getattr(profile, "user_type", "") if profile else ""

    if program.status == "hidden":
        if request.user.is_authenticated:
            if request.user.is_staff:
                pass
            elif user_type == "center_teacher" and profile and program.teacher == profile:
                pass
            else:
                messages.error(request, "비공개 프로그램입니다.")
                return redirect("program_list")
        else:
            messages.error(request, "비공개 프로그램입니다.")
            return redirect("program_list")

    can_apply = (
        request.user.is_authenticated
        and (request.user.is_superuser or user_type in ["parent", "student"])
    )

    # ✅ 핵심: 반 목록을 직접 만들고, 필요한 데이터 붙이기
    classes = (
        program.classes
        .all()
        .prefetch_related("applications", "enrollments")
    )

    for cls in classes:
        # 🔹 신청접수만 카운트
        cls.pending_applications = cls.applications.filter(status="pending")

        # 🔹 실제 수강중인 학생만
        cls.active_enrollments = cls.enrollments.filter(is_active=True)\
            .select_related("student", "student__user")

    return render(request, "courses/program_detail.html", {
        "program": program,
        "can_apply": can_apply,
        "classes": classes,  # ❗ program.classes 쓰지 말 것
    })





@login_required
def program_apply(request, pk):
    program = get_object_or_404(Program, pk=pk)
    profile = request.user.profile
    user_type = getattr(profile, "user_type", "")

    if user_type not in ["parent", "student"] and not request.user.is_superuser:
        messages.error(request, "학부모/학생만 신청할 수 있습니다.")
        return redirect("program_detail", pk=pk)

    if request.method == "POST":
        form = ProgramApplicationForm(request.POST)

        # ✅ 반 선택 필수
        class_id = request.POST.get("class_id")
        try:
            selected_class = ProgramClass.objects.get(id=class_id, program=program)
        except (ProgramClass.DoesNotExist, TypeError, ValueError):
            messages.error(request, "반을 선택해주세요.")
            return redirect("program_apply", pk=program.id)

        if user_type == "parent":
            # 부모는 자녀 선택 필수
            child_ids = request.POST.getlist("children")
            if not child_ids:
                messages.error(request, "수강할 자녀를 선택해주세요.")
                return redirect("program_apply", pk=program.id)

            memo = form.cleaned_data.get("memo", "") if form.is_valid() else ""

            for cid in child_ids:
                try:
                    child = Child.objects.get(id=cid, parent=profile)
                    
                    # ✅ 대상 나이 검증
                    if not is_child_in_target(child, program):
                        messages.error(
                            request,
                            f"{child.name} (만 {child.birth_date.year}년생)은 이 프로그램 대상 연령이 아닙니다."
                        )
                        return redirect("program_apply", pk=pk)
                    
                    # 중복 신청 방지 (같은 프로그램+반+자녀 조합 방지)
                    if not ProgramApplication.objects.filter(program=program, program_class=selected_class, child=child).exists():
                        ProgramApplication.objects.create(
                            program=program,
                            program_class=selected_class,   # ✅ 선택된 반 저장
                            applicant=profile,
                            child=child,
                            applicant_name=profile.user.first_name or profile.user.username,
                            phone=profile.phone_number,
                            memo=memo,
                        )
                except Child.DoesNotExist:
                    continue

            messages.success(request, "자녀 수강신청이 접수되었습니다. (관리자 승인 대기)")
            return redirect("program_detail", pk=program.id)

        else:  # 학생 본인 신청
            if form.is_valid():
                memo = form.cleaned_data.get("memo", "")
                if not ProgramApplication.objects.filter(program=program, program_class=selected_class, applicant=profile, child__isnull=True).exists():
                    ProgramApplication.objects.create(
                        program=program,
                        program_class=selected_class,   # ✅ 선택된 반 저장
                        applicant=profile,
                        child=None,
                        applicant_name=profile.user.first_name,
                        phone=profile.phone_number,
                        memo=memo,
                    )
                messages.success(request, "수강신청이 접수되었습니다. (관리자 승인 대기)")
                return redirect("program_detail", pk=program.id)
            else:
                messages.error(request, "입력값을 확인해주세요.")

    else:
        form = ProgramApplicationForm()

    children = []
    if user_type == "parent":
        children = profile.children.all()

    return render(request, "courses/program_apply.html", {
        "program": program,
        "form": form,
        "profile": profile,
        "children": children,
    })



@login_required
@user_passes_test(lambda u: u.is_staff)  # 관리자만
def program_delete(request, pk):
    program = get_object_or_404(Program, pk=pk)
    if request.method == "POST":
        program.delete()
        messages.success(request, "프로그램이 삭제되었습니다.")
        return redirect('program_list')
    return render(request, "courses/program_confirm_delete.html", {"program": program})


#프로그램 예약
@login_required
def reservation_create(request):
    """기관 예약 생성"""
    if request.method == "POST":
        form = InstitutionReservationForm(request.POST, user=request.user)
        if form.is_valid():
            reservation = form.save(commit=False)
            reservation.requested_by = request.user

            # ✅ 기관 로그인인 경우 institution 자동 지정
            if hasattr(request.user, "profile") and request.user.profile.user_type == "institution":
                reservation.institution = request.user

            reservation.save()
            messages.success(request, "예약 요청이 완료되었습니다.")
            return redirect("reservation_list")
        else:
            # ✅ 어떤 필드가 문제인지 표시
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"[{form.fields[field].label}] {error}")
    else:
        form = InstitutionReservationForm(user=request.user)
        
    return render(request, "product/reservation_form.html", {"form": form})

# ✅ 예약 목록 뷰
@login_required
def reservation_list(request):
    """
    예약 목록 출력 뷰
    - 관리자: 전체 예약 조회
    - 기관 사용자: 자기 User 계정으로 등록한 예약만 조회
    """
    if request.user.is_staff:  # 관리자
        reservations = InstitutionReservation.objects.all().select_related("institution", "product")
    else:  # 기관 사용자
        reservations = InstitutionReservation.objects.filter(
            institution=request.user
        ).select_related("institution", "product")

    return render(request, "product/reservation_list.html", {"reservations": reservations})

# 예약 확정/확정취소 토글
@login_required
def reservation_confirm(request, pk):
    """
    관리자 전용 예약 확정/확정취소 토글
    - requested → approved
    - approved → requested
    """
    reservation = get_object_or_404(InstitutionReservation, pk=pk)

    if not request.user.is_staff:
        messages.error(request, "승인 권한이 없습니다.")
        return redirect("reservation_list")

    if reservation.status == "approved":
        reservation.status = "requested"
        messages.success(request, f"[{reservation.product.name}] 예약이 '예약요청' 상태로 변경되었습니다.")
    else:
        reservation.status = "approved"
        messages.success(request, f"[{reservation.product.name}] 예약이 '예약확정' 상태로 변경되었습니다.")

    reservation.save()
    return redirect("reservation_list")

#예약 취소
@login_required
def reservation_cancel(request, pk):
    reservation = get_object_or_404(InstitutionReservation, pk=pk)
    # 관리자: 삭제, 기관: 상태를 'canceled' 로만 변경
    if request.user.is_staff:
        reservation.delete()
        messages.success(request, "예약이 삭제되었습니다.")
    else:
        reservation.status = "canceled"
        reservation.save()
        messages.success(request, "예약이 취소되었습니다.")
    return redirect("reservation_list")

#예약 수정
@login_required
def reservation_edit(request, pk):
    reservation = get_object_or_404(InstitutionReservation, pk=pk)

    # 권한 체크: 관리자는 모두 수정 가능, 기관은 자기 것만
    if not request.user.is_staff and reservation.institution != request.user:
        messages.error(request, "해당 예약을 수정할 권한이 없습니다.")
        return redirect("reservation_list")

    if request.method == "POST":
        form = InstitutionReservationForm(request.POST, instance=reservation)
        if form.is_valid():
            form.save()
            messages.success(request, "예약이 수정되었습니다.")
            return redirect("reservation_list")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"[{form.fields[field].label}] {error}")
    else:
        form = InstitutionReservationForm(instance=reservation)

    return render(
        request,
        "product/reservation_form.html",
        {
            "form": form,
            "reservation": reservation,  # ✅ 템플릿에서 사용 가능하게 넘기기
        },
    )

@login_required
def reservation_calendar(request):
    """예약 달력 페이지"""
    return render(request, "product/reservation_calendar.html")


@login_required
def reservation_events(request):
    """FullCalendar 이벤트 JSON"""
    reservations = InstitutionReservation.objects.select_related(
        "institution", "product", "institution__institution_profile"
    ).all()

    events = []
    for r in reservations:
        institution_name = (
            r.institution.institution_profile.institution_name
            if hasattr(r.institution, "institution_profile") else r.institution.username
        )

        events.append({
            "title": r.product.name,   # 프로그램명
            "start": f"{r.date}T{r.start_time}",
            "end": f"{r.date}T{r.end_time}" if r.end_time else None,
            "extendedProps": {
                "institution": institution_name,   # ✅ 기관 이름 우선, 없으면 username
                "status": r.get_status_display(),
                "status_code": r.status,
                "place": r.place,
                "headcount": r.headcount,
                "topic": r.selected_topic,  # ✅ 주제 추가
            }
        })
    return JsonResponse(events, safe=False)



# --- 권한: 관리자/스태프만 ---
def _is_staff(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

# 목록/상세 (누구나 로그인 시 열람)
@login_required
def product_list(request):
    category_id = request.GET.get("category")
    categories = Category.objects.all()

    # ✅ 관리자와 일반 사용자 분리
    if request.user.is_staff:
        products = ProgramProduct.objects.all().order_by("-created_at")
    else:
        products = ProgramProduct.objects.filter(status="public").order_by("-created_at")

    selected_category = None
    if category_id:
        try:
            selected_category = int(category_id)
            products = products.filter(category_id=selected_category)
        except ValueError:
            pass

    return render(request, "product/product_list.html", {
        "products": products,
        "categories": categories,
        "selected_category": selected_category,
    })

@login_required
def product_detail(request, pk):
    product = get_object_or_404(ProgramProduct, pk=pk)
    materials = product.materials.all()

    # ✅ 재료 합계
    material_total = sum(m.price for m in materials)

    # ✅ 총 수강료 (포함이면 = base_price, 미포함이면 = base_price + 재료비 합계)
    if product.include_material_cost:
        total_price = product.base_price
    else:
        total_price = product.base_price + material_total

    return render(
        request,
        "product/product_detail.html",
        {
            "product": product,
            "material_total": material_total,
            "total_price": total_price,
        }
    )

def product_detail_api(request, pk):
    product = ProgramProduct.objects.get(pk=pk)
    data = {
        "id": product.id,
        "name": product.name,
        "base_price": product.base_price,
        "duration_minutes": product.duration_minutes,
        "category": product.category.name if product.category else None,
        "topics": product.topics or [],
        "include_material_cost": product.include_material_cost,
        "included_materials": product.included_materials,
        "materials": [
            {"name": m.name, "price": m.price}
            for m in product.materials.all()
        ]
    }
    return JsonResponse(data)


# 생성
@login_required
@user_passes_test(_is_staff)
def product_create(request):
    if request.method == "POST":
        print("📌 request.FILES:", request.FILES)  # 업로드된 파일 확인
        form = ProgramProductForm(request.POST, request.FILES)
        formset = ProductMaterialFormSet(request.POST, request.FILES)  # ✅ FILES 전달

        if form.is_valid() and formset.is_valid():
            product = form.save()
            formset.instance = product
            formset.save()
            print("📌 저장된 image 필드:", product.image, product.image.url if product.image else None)
            messages.success(request, "프로그램이 등록되었습니다.")
            return redirect("product_list")
        else:
            print("📌 form.errors:", form.errors.as_json())
            print("📌 formset.errors:", [fs.errors for fs in formset])
    else:
        form = ProgramProductForm()
        formset = ProductMaterialFormSet()

    return render(
        request,
        "product/product_form.html",
        {"form": form, "formset": formset, "mode": "create"},
    )


# 수정
@login_required
@user_passes_test(_is_staff)
def product_update(request, pk):
    product = get_object_or_404(ProgramProduct, pk=pk)

    if request.method == "POST":
        form = ProgramProductForm(request.POST, request.FILES, instance=product)
        formset = ProductMaterialFormSet(request.POST, instance=product)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "프로그램이 수정되었습니다.")
            return redirect("product_list")
        else:
            print("📌 form.errors:", form.errors)
            print("📌 formset.errors:", formset.errors)
    else:
        form = ProgramProductForm(instance=product)
        formset = ProductMaterialFormSet(instance=product)

    return render(
        request,
        "product/product_form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "update",
            "product": product,
        },
    )


# 삭제(POST 전용)
@login_required
@user_passes_test(_is_staff)
def product_delete(request, pk):
    product = get_object_or_404(ProgramProduct, pk=pk)
    if request.method != "POST":
        return HttpResponseForbidden("잘못된 접근입니다.")
    product.delete()
    messages.success(request, "프로그램이 삭제되었습니다.")
    return redirect("product_list")

def category_list(request):
    categories = Category.objects.all()
    return render(request, "product/category_list.html", {"categories": categories})

def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "카테고리가 추가되었습니다.")
            return redirect("category_list")
    else:
        form = CategoryForm()
    return render(request, "product/category_form.html", {"form": form, "mode": "create"})

def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "카테고리가 수정되었습니다.")
            return redirect("category_list")
    else:
        form = CategoryForm(instance=category)
    return render(request, "product/category_form.html", {"form": form, "mode": "update"})

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        messages.success(request, "카테고리가 삭제되었습니다.")
        return redirect("category_list")
    return redirect("category_list")

def product_api(request, pk):
    """프로그램 상세 JSON 응답"""
    try:
        product = get_object_or_404(ProgramProduct, pk=pk)
    except ProgramProduct.DoesNotExist:
        raise Http404("Product not found")

    data = {
        "id": product.id,
        "name": product.name,
        "base_price": product.base_price,
        "duration_minutes": product.duration_minutes,
        "category": str(product.category) if product.category else None,
        "topics": product.topics or [],
    }
    return JsonResponse(data)

@login_required
def program_clone(request, pk):
    original = get_object_or_404(Program, pk=pk)

    if request.method == "POST":
        form = ProgramForm(request.POST, request.FILES)
        if form.is_valid():
            cloned = form.save(commit=False)
            cloned.pk = None  # ✅ 새 객체로 저장
            if not cloned.image and original.image:
                cloned.image = original.image  # ✅ 이미지도 복사
            cloned.save()
            form.save_m2m()
            return redirect("program_detail", pk=cloned.pk)
    else:
        # ✅ 기존 값 그대로 초기화
        initial = {
            "name": f"{original.name} (복제본)",

            # 👉 target_start / target_end로 교체
            "target_start": original.target_start_id,
            "target_end": original.target_end_id,

            "teacher": original.teacher_id,
            "recruit_start_date": original.recruit_start_date,
            "recruit_end_date": original.recruit_end_date,
            "start_date": original.start_date,
            "end_date": original.end_date,

            # class_days는 이미 Program에 문자열 필드로 있다고 가정
            "class_days": original.class_days.split(",") if getattr(original, "class_days", None) else [],

            "class_time_start": getattr(original, "class_time_start", None),
            "class_time_end": getattr(original, "class_time_end", None),

            "session_count": original.session_count,
            "tuition": f"{original.tuition:,}" if original.tuition else "120,000",
            "status": original.status,
            "description": original.description,
        }
        form = ProgramForm(initial=initial)

    return render(request, "courses/program_form.html", {"form": form, "program": None})
from .models import Target
from .forms import TargetForm

# ✅ 관리자만 접근 가능
def is_admin(user):
    return user.is_staff

@login_required
@user_passes_test(is_admin)
def target_list(request):
    targets = Target.objects.all().order_by("id")
    return render(request, "courses/target_list.html", {"targets": targets})

@login_required
@user_passes_test(is_admin)
def target_create(request):
    if request.method == "POST":
        form = TargetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "대상이 추가되었습니다.")
            return redirect("target_list")
    else:
        form = TargetForm()
    return render(request, "courses/target_form.html", {"form": form, "mode": "create"})

@login_required
@user_passes_test(is_admin)
def target_update(request, pk):
    target = get_object_or_404(Target, pk=pk)
    if request.method == "POST":
        form = TargetForm(request.POST, instance=target)
        if form.is_valid():
            form.save()
            messages.success(request, "대상이 수정되었습니다.")
            return redirect("target_list")
    else:
        form = TargetForm(instance=target)
    return render(request, "courses/target_form.html", {"form": form, "mode": "update"})

@login_required
@user_passes_test(is_admin)
def target_delete(request, pk):
    target = get_object_or_404(Target, pk=pk)
    if request.method == "POST":
        target.delete()
        messages.success(request, "대상이 삭제되었습니다.")
        return redirect("target_list")
    return render(request, "courses/target_confirm_delete.html", {"target": target})

@login_required
def student_course_list(request):
    programs = LearningProgram.objects.all().order_by("id")

    enrolled = LearningEnrollment.objects.filter(user=request.user)\
                                        .values_list("program__id", flat=True)

    return render(request, "courses/student_course_list.html", {
        "programs": programs,
        "enrolled": enrolled,
    })


@login_required
def student_course_apply(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)

    # 중복 체크
    if LearningEnrollment.objects.filter(user=request.user, program=program).exists():
        messages.warning(request, "이미 신청한 프로그램입니다.")
        return redirect("student_course_list")

    LearningEnrollment.objects.create(user=request.user, program=program)
    messages.success(request, f"{program.name} 수강신청 완료!")

    # 신청 후 해당 프로그램 URL로 이동
    return redirect(program.get_url())


@login_required
@user_passes_test(is_admin)
def create_learning_program(request):
    if request.method == "POST":
        code = request.POST.get("code")
        name = request.POST.get("name")
        desc = request.POST.get("description")
        image = request.FILES.get("image")

        LearningProgram.objects.create(
            code=code,
            name=name,
            description=desc,
            image=image
        )
        messages.success(request, "새 LearningProgram이 추가되었습니다.")
        return redirect("student_course_list")

    return render(request, "courses/learning_program_form.html")

# ================================
# 📘 LearningProgram 관리자 CRUD
# ================================

# ▶ 목록
@login_required
@user_passes_test(is_admin)
def learning_program_list(request):
    programs = LearningProgram.objects.all().order_by("id")
    return render(request, "courses/learning_program_list.html", {
        "programs": programs
    })

# ▶ 생성
@login_required
@user_passes_test(is_admin)
def learning_program_create(request):
    if request.method == "POST":
        form = LearningProgramForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "새 학습 프로그램이 추가되었습니다.")
            return redirect("learning_program_list")
    else:
        form = LearningProgramForm()

    return render(request, "courses/learning_program_form.html", {
        "form": form,
        "mode": "create",
    })

# ▶ 수정
@login_required
@user_passes_test(is_admin)
def learning_program_edit(request, pk):
    program = get_object_or_404(LearningProgram, pk=pk)

    if request.method == "POST":
        form = LearningProgramForm(request.POST, request.FILES, instance=program)
        if form.is_valid():
            form.save()
            messages.success(request, "학습 프로그램이 수정되었습니다.")
            return redirect("learning_program_list")
    else:
        form = LearningProgramForm(instance=program)

    return render(request, "courses/learning_program_form.html", {
        "form": form,
        "program": program,
        "mode": "edit",
    })

# ▶ 삭제
@login_required
@user_passes_test(is_admin)
def learning_program_delete(request, pk):
    program = get_object_or_404(LearningProgram, pk=pk)

    if request.method == "POST":
        program.delete()
        messages.success(request, "학습 프로그램이 삭제되었습니다.")
        return redirect("learning_program_list")

    return render(request, "courses/learning_program_delete.html", {
        "program": program
    })

def chapter_manage(request, pk):
    program = get_object_or_404(LearningProgram, id=pk)

    if request.method == "POST" and "excel_file" in request.FILES:
        file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(file)

        # 기존 데이터 삭제
        Chapter.objects.filter(program=program).delete()

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            for row in ws.iter_rows(min_row=2, values_only=True):

                # 빈 줄 무시
                if not any(row):
                    continue

                # row 길이 맞추기 (10개)
                row = list(row)
                if len(row) < 10:
                    row += [None] * (10 - len(row))
                row = row[:10]

                (
                    chapter_no,
                    chapter_title,
                    chapter_content,
                    key,
                    title,
                    item_type,
                    explain_html,
                    hint,
                    answer_code,
                    expected_output
                ) = row

                # chapter_no 또는 chapter_title이 없으면 skip
                if not chapter_no and not chapter_title:
                    continue

                # 챕터 생성 또는 가져오기
                chapter, created = Chapter.objects.get_or_create(
                    program=program,
                    number=chapter_no,
                    defaults={
                        "title": chapter_title or f"{chapter_no}장",
                        "content": chapter_content
                    }
                )

                # 기존 챕터 업데이트 처리 (중요!)
                if not created:
                    chapter.title = chapter_title or chapter.title
                    chapter.content = chapter_content or chapter.content
                    chapter.save()

                # title이 없으면 기본값 처리
                item_title = title or key or "제목없음"

                # 아이템 생성
                Item.objects.create(
                    chapter=chapter,
                    key=key,
                    title=item_title,
                    item_type=item_type,
                    explain_html=explain_html,
                    hint=hint,
                    answer_code=answer_code,
                    expected_output=expected_output
                )

        messages.success(request, "엑셀 업로드 완료!")
        return redirect("chapter_manage", pk=program.id)

    chapters = Chapter.objects.filter(program=program).order_by("number")

    return render(request, "courses/chapter_manage.html", {
        "program": program,
        "chapters": chapters
    })


def course_home(request, program_id):

    program = get_object_or_404(LearningProgram, id=program_id)

    chapters = Chapter.objects.filter(
        program=program
    ).order_by("number")

    progress = {}
    if request.user.is_authenticated:
        for ch in chapters:
            total = Item.objects.filter(chapter=ch).count()
            completed = UserProgress.objects.filter(
                user=request.user,
                item__chapter=ch,
                completed=True
            ).count()

            progress[ch.id] = round((completed / total) * 100) if total else 0

    return render(request, "learning_program/course_home.html", {
        "program": program,
        "chapters": chapters,
        "progress": progress
    })



def chapter_detail(request, chapter_id):
    chapter = get_object_or_404(Chapter, id=chapter_id)

    # 🔥 같은 프로그램 소속 아이템만 로딩
    items = Item.objects.filter(chapter=chapter).order_by("key")

    # 🔥 1) item별 완료 여부 dict
    user_progress = {}
    if request.user.is_authenticated:
        progresses = UserProgress.objects.filter(
            user=request.user,
            item__in=items
        )
        for up in progresses:
            user_progress[up.item.id] = up.completed

    # 🔥 2) 챕터 전체 진도율
    total = items.count()
    completed_count = sum(1 for x in user_progress.values() if x)
    chapter_progress = round(completed_count / total * 100) if total else 0

    return render(request, "learning_program/chapter_detail.html", {
        "chapter": chapter,
        "items": items,
        "user_progress": user_progress,
        "chapter_progress": chapter_progress,
        "program": chapter.program,   # 🔥 템플릿에서 프로그램 정보 사용 가능
    })


def item_page(request, item_id):
    item = get_object_or_404(Item, id=item_id)

    # 🔥 같은 chapter 내 item 전체 목록 (id 기준 정렬)
    items = list(Item.objects.filter(chapter=item.chapter).order_by("id"))

    # 🔥 현재 item의 index 찾기
    idx = items.index(item)

    # 🔥 이전/다음 item 계산
    prev_item = items[idx - 1] if idx > 0 else None
    next_item = items[idx + 1] if idx < len(items) - 1 else None

    # ============================================================
    # 🔥 프로그램 이름 기반 템플릿 분기 (program_type → program.name)
    # ============================================================
    program_type = None

    program = item.chapter.program
    if program and program.name:
        program_type = program.name.lower().strip()
    else:
        program_type = ""  # None 방지

    print("🔥 program_type:", program_type)

    # ============================================================
    # 🔥 ITQ 파워포인트일 경우 OneDrive URL 생성
    # ============================================================
    onedrive_url = None

    if "itq파워포인트" in program_type or "ppt" in program_type:
        folder_name = item.chapter.title.strip()
        file_name = f"{item.key}.pptx"

        BASE_ONEDRIVE_URL = (
            "https://makinglab-my.sharepoint.com/:f:/g/personal/kly120112_steam-making_com/IgD_WziHDDRMSoqFfD6VSxg3AWLou17QyWWG2y6ekKKryAE?e=Abgyeu"
        )

        onedrive_url = f"{BASE_ONEDRIVE_URL}"
        print(onedrive_url)

    # 🔥 템플릿 매핑 (program.name 기준)
    template_map = {
        "파이썬": "learning_program/item_page_python.html",
        "python": "learning_program/item_page_python.html",
        "itq파워포인트": "learning_program/item_page_itq_powerpoint.html",
        "itq엑셀": "learning_program/item_page_itq_excel.html",
        "로봇": "learning_program/item_page_robot.html",
        "ai": "learning_program/item_page_ai.html",
    }

    # 🔥 매칭되는 템플릿이 없으면 기본 템플릿 사용
    template_name = "learning_program/item_page.html"
    for key, tpl in template_map.items():
        if key in program_type:
            template_name = tpl
            break

    # ============================================================
    # 🔥 최종 렌더링
    # ============================================================
    return render(request, template_name, {
        "item": item,
        "prev_item": prev_item,
        "next_item": next_item,
        "onedrive_url": onedrive_url,
    })



def update_progress(request, item_id):
    if request.method == "POST":
        status = request.POST.get("status")

        progress, created = UserProgress.objects.get_or_create(
            user=request.user,
            item_id=item_id
        )

        if status == "done":
            progress.completed = True
        elif status == "hold":
            progress.completed = False  # 또는 상태값 확장 가능

        progress.save()

    return redirect("item_page", item_id=item_id)


def run_code(request):
    code = request.POST.get("code", "")
    input_raw = request.POST.get("input_value", "")

    # 여러 줄 input 분리
    input_lines = input_raw.split("\n")

    # input() override
    fake_input = "input_values = []\n"
    for line in input_lines:
        fake_input += f"input_values.append('{line}')\n"

    fake_input += """
_input_index = 0
def input(prompt=None):
    global _input_index
    if _input_index < len(input_values):
        value = input_values[_input_index]
        _input_index += 1
        return value
    return ""
"""

    exec_code = fake_input + "\n" + code

    # print 캡처
    old_stdout = sys.stdout
    sys.stdout = io.StringIO()

    try:
        exec(exec_code, {})
        output = sys.stdout.getvalue()
    except Exception as e:
        output = str(e)

    sys.stdout = old_stdout

    return JsonResponse({"output": output})




def grade_code(request):
    item_id = request.POST['item_id']
    code = request.POST['code']

    item = Item.objects.get(id=item_id)
    output = safe_exec(code)

    # 🔍 출력 비교: strip() + 양쪽 공백/줄바꿈 제거
    user_out = output.strip()
    expected_out = item.expected_output.strip()

    score = 100 if user_out == expected_out else 0

    # 🔥 사용자 기록 저장
    if request.user.is_authenticated:
        progress, created = UserProgress.objects.get_or_create(
            user=request.user, item=item
        )
        progress.code = code
        progress.last_output = output
        progress.score = score
        progress.completed = (score == 100)
        progress.save()

    return JsonResponse({
        "score": score,
        "output": output,
        "expected": item.expected_output,
        "completed": (score == 100)   # 🔥 프런트에서 사용
    })



def get_hint(request, item_id):
    item = Item.objects.get(id=item_id)
    return JsonResponse({"hint": item.hint})


def get_answer(request, item_id):
    item = Item.objects.get(id=item_id)
    return JsonResponse({"answer": item.answer_code})


def program_curriculum(request, program_id):
    program = get_object_or_404(Program, id=program_id)
    syllabus = program.syllabus.all()

    return render(request, "courses/curriculum/program_curriculum.html", {
        "program": program,
        "syllabus": syllabus,
    })

@staff_member_required
def curriculum_program_create(request):
    if request.method == "POST":
        form = CurriculumProgramForm(request.POST)
        if form.is_valid():
            program = form.save()
            return redirect("curriculum_program_list")
    else:
        form = CurriculumProgramForm()

    return render(
        request,
        "courses/curriculum/program_curriculum_form.html",
        {
            "form": form,
            "mode": "create",
        }
    )

@staff_member_required
def curriculum_program_update(request, program_id):
    program = get_object_or_404(CurriculumProgram, id=program_id)

    if request.method == "POST":
        form = CurriculumProgramForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            return redirect("curriculum_program_list")
    else:
        form = CurriculumProgramForm(instance=program)

    return render(
        request,
        "courses/curriculum/program_curriculum_form.html",
        {
            "form": form,
            "mode": "edit",
            "program": program,
        }
    )

@staff_member_required
def curriculum_program_delete(request, program_id):
    program = get_object_or_404(CurriculumProgram, id=program_id)
    program.delete()
    return redirect("curriculum_program_list")


@user_passes_test(is_admin)
def program_curriculum_edit(request, program_id):
    program = get_object_or_404(Program, id=program_id)

    if request.method == "POST":
        form = ProgramForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            return redirect(
                "program_curriculum",
                program_id=program.id
            )
    else:
        form = ProgramForm(instance=program)

    return render(
        request,
        "courses/curriculum/program_curriculum_edit.html",  # ⭐ 분리
        {
            "program": program,
            "form": form,
        }
    )


def syllabus_upload(request, program_id):
    program = get_object_or_404(Program, id=program_id)

    if request.method == "POST":
        form = SyllabusUploadForm(request.POST, request.FILES)
        if form.is_valid():
            import_syllabus_from_excel(
                program,
                request.FILES["excel_file"]
            )
            return redirect(
                "program_curriculum",
                program_id=program.id
            )
    else:
        form = SyllabusUploadForm()

    return render(
        request,
        "courses/curriculum/syllabus_upload.html",
        {
            "program": program,
            "form": form,
        }
    )

@login_required
def curriculum_program_list(request):
    programs = CurriculumProgram.objects.all()

    return render(
        request,
        "courses/curriculum/program_curriculum_list.html",
        {
            "programs": programs,
        }
    )

@staff_member_required
def curriculum_syllabus_create(request, program_id):
    program = get_object_or_404(CurriculumProgram, id=program_id)

    if request.method == "POST":
        form = CurriculumSyllabusForm(request.POST)
        if form.is_valid():
            syllabus = form.save(commit=False)
            syllabus.program = program
            syllabus.save()
            return redirect(
                "curriculum_syllabus_list",
                program_id=program.id
            )
    else:
        form = CurriculumSyllabusForm()

    return render(
        request,
        "courses/curriculum/program_curriculum_syllabus_form.html",
        {
            "form": form,
            "program": program,
            "mode": "create",  # or "edit"
        }
    )

@staff_member_required
def curriculum_syllabus_update(request, syllabus_id):
    syllabus = get_object_or_404(CurriculumSyllabus, id=syllabus_id)
    program = syllabus.program  # ✅ 이 줄이 핵심

    if request.method == "POST":
        form = CurriculumSyllabusForm(request.POST, instance=syllabus)
        if form.is_valid():
            form.save()
            return redirect(
                "curriculum_syllabus_list",
                program_id=program.id
            )
    else:
        form = CurriculumSyllabusForm(instance=syllabus)

    return render(
        request,
        "courses/curriculum/program_curriculum_syllabus_form.html",
        {
            "form": form,
            "program": program,   # ✅ 이제 정상
            "mode": "edit"
        }
    )


@staff_member_required
def curriculum_syllabus_delete(request, syllabus_id):
    syllabus = get_object_or_404(CurriculumSyllabus, id=syllabus_id)
    program_id = syllabus.program.id
    syllabus.delete()

    return redirect(
        "curriculum_syllabus_list",
        program_id=program_id
    )

@login_required
def curriculum_syllabus_list(request, program_id):
    program = get_object_or_404(CurriculumProgram, id=program_id)
    syllabus_list = program.syllabus.all()

    return render(
        request,
        "courses/curriculum/program_curriculum_manage.html",
        {
            "program": program,
            "syllabus_list": syllabus_list,
        }
    )

import openpyxl
from django.contrib import messages

@staff_member_required
def curriculum_syllabus_excel_upload(request, program_id):
    program = get_object_or_404(CurriculumProgram, id=program_id)

    if request.method == "POST":
        form = CurriculumSyllabusExcelForm(request.POST, request.FILES)
        if form.is_valid():
            wb = openpyxl.load_workbook(request.FILES["excel_file"])
            ws = wb.active

            # 2행부터 데이터 (1행은 헤더)
            created_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                week, title, content, material = row

                if not week or not title:
                    continue

                CurriculumSyllabus.objects.create(
                    program=program,
                    week=int(week),
                    title=str(title),
                    content=str(content or ""),
                    material=str(material or ""),
                )
                created_count += 1

            messages.success(
                request,
                f"{created_count}개의 차시가 등록되었습니다."
            )

            return redirect(
                "curriculum_syllabus_list",
                program_id=program.id
            )
    else:
        form = CurriculumSyllabusExcelForm()

    return render(
        request,
        "courses/curriculum/program_curriculum_excel_upload.html",
        {
            "form": form,
            "program": program,
        }
    )

import openpyxl
from django.http import HttpResponse
from datetime import date


@staff_member_required
def curriculum_syllabus_excel_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "차시 템플릿"

    # ✅ 헤더
    headers = ["week", "title", "content", "material"]
    ws.append(headers)

    # ✅ 예시 데이터
    ws.append([1, "로봇의 기본 구조 이해", "로봇의 구성 요소를 알아봅니다.", "로봇 키트"])
    ws.append([2, "모터 제어 실습", "모터의 회전 방향을 제어합니다.", "모터, 배터리"])

    # 컬럼 너비 자동 조정
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25

    # 응답 생성
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"curriculum_syllabus_template_{date.today()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from accounts.utils import create_student_account
from .models import ProgramApplication, ProgramEnrollment


@staff_member_required
def convert_to_enrollment(request, app_id):
    app = get_object_or_404(ProgramApplication, id=app_id)

    # ✅ 1️⃣ 이미 승인된 신청이면 차단
    if app.status == "approved":
        messages.warning(request, "이미 승인된 신청입니다.")
        return redirect("program_detail", pk=app.program.id)

    # ❌ 자녀 없는 신청은 승인 불가
    if not app.child:
        messages.error(request, "자녀 정보가 없는 신청은 승인할 수 없습니다.")
        return redirect("program_detail", pk=app.program.id)

    child = app.child
    program = app.program
    program_class = app.program_class

    # ✅ 2️⃣ 학생 계정 생성 or 재사용
    student_profile = create_student_account(child)

    # ✅ 3️⃣ 중복 수강 방지 (DB + 로직 이중 보호)
    if ProgramEnrollment.objects.filter(
        program_class=program_class,
        student=student_profile,
        is_active=True
    ).exists():
        messages.warning(
            request,
            f"{child.name} 학생은 이미 수강 중입니다."
        )
        return redirect("program_detail", pk=program.id)

    # ✅ 4️⃣ 수강생 등록
    ProgramEnrollment.objects.create(
        program=program,
        program_class=program_class,
        student=student_profile
    )

    # ✅ 5️⃣ 신청 → 승인 처리
    app.status = "approved"
    app.save(update_fields=["status"])

    #app.delete()

    messages.success(
        request,
        f"{child.name} 학생 계정 생성 및 수강 등록이 완료되었습니다."
    )
    return redirect("program_detail", pk=program.id)



@staff_member_required
def program_enrollment_add_global(request, program_id):
    if request.method != "POST":
        return redirect("program_detail", pk=program_id)

    program = get_object_or_404(Program, id=program_id)

    class_id = request.POST.get("class_id")
    student_id = request.POST.get("student_id")

    if not class_id or not student_id:
        messages.error(request, "반과 학생을 모두 선택해주세요.")
        return redirect("program_detail", pk=program_id)

    program_class = get_object_or_404(
        ProgramClass,
        id=class_id,
        program=program
    )

    # ✅ 학생 = 회원 Profile
    student = get_object_or_404(
        Profile,
        id=student_id,
        user_type="student"
    )

    # ✅ 중복 등록 방지
    if ProgramEnrollment.objects.filter(
        program_class=program_class,
        student=student,
        is_active=True
    ).exists():
        messages.warning(
            request,
            f"{student.user.get_full_name() or student.user.username} 님은 이미 등록되어 있습니다."
        )
        return redirect("program_detail", pk=program_id)

    # ✅ 수강생 등록
    ProgramEnrollment.objects.create(
        program=program,
        program_class=program_class,
        student=student
    )

    messages.success(
        request,
        f"{student.user.get_full_name() or student.user.username} 님이 수강생으로 등록되었습니다."
    )
    return redirect("program_detail", pk=program_id)

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from courses.models import ProgramEnrollment, ProgramApplication

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

from courses.models import ProgramEnrollment, ProgramApplication
from accounts.models import Child


@staff_member_required
def cancel_enrollment(request, enrollment_id):
    enrollment = get_object_or_404(ProgramEnrollment, id=enrollment_id)

    program = enrollment.program
    program_class = enrollment.program_class
    student_profile = enrollment.student

    # 🔹 1. 해당 학생의 Child 찾기
    child = Child.objects.filter(
        student_profile=student_profile
    ).first()

    # 🔹 2. 연결된 신청 삭제
    if child:
        ProgramApplication.objects.filter(
            program=program,
            program_class=program_class,
            child=child
        ).delete()

    # 🔹 3. 수강생 삭제
    enrollment.delete()

    messages.success(request, "수강이 취소되었습니다.")
    return redirect("program_detail", pk=program.id)




@staff_member_required
def reject_application(request, app_id):
    app = get_object_or_404(ProgramApplication, id=app_id)
    app.status = "rejected"
    app.save()
    return redirect("program_detail", app.program.id)

@staff_member_required
def delete_application(request, app_id):
    app = get_object_or_404(ProgramApplication, id=app_id)
    program_id = app.program.id
    app.delete()
    return redirect("program_detail", program_id)

from django.http import JsonResponse
from django.db.models import Q
from accounts.models import Profile
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def search_members(request):
    q = request.GET.get("q", "").strip()
    user_type = request.GET.get("user_type", "").strip()

    profiles = Profile.objects.select_related("user")

    # ✅ 회원유형 필터
    if user_type:
        profiles = profiles.filter(user_type=user_type)

    # ✅ 검색어
    if q:
        profiles = profiles.filter(
            Q(user__first_name__icontains=q) |
            Q(user__username__icontains=q)
        )

    data = []
    for p in profiles[:20]:
        data.append({
            "id": p.id,   # ⭐ Profile.id
            "name": p.user.get_full_name() or p.user.username,
            "user_type": p.user_type,
        })

    return JsonResponse(data, safe=False)

