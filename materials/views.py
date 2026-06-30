
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Material, Vendor, VendorType, Order, Receipt, MaterialOrder, MaterialOrderItem
from teachers.models import TeachingInstitution
from django.contrib.auth.models import User
from .forms import MaterialForm, VendorForm, VendorTypeForm, OrderForm, ReceiptForm
from django.db.models import Q, Case, When, Value, IntegerField, Exists, OuterRef
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST
from .models import MaterialRelease, MaterialReleaseItem  # ✅ 출고 모델 import 확인
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Case, When, Value, IntegerField
from django.db.models.functions import ExtractYear
from django.shortcuts import render
from collections import defaultdict
from django.db.models import Sum, F
from io import BytesIO
from django.http import FileResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.db.models import Exists, OuterRef
from django.urls import reverse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer

from .models import MaterialRelease, MaterialReleaseItem, ReleasePaymentStatus
from teachers.models import TeachingInstitution
from django.contrib.auth import get_user_model
# ✅ 순서 업데이트 (AJAX)
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.forms import modelformset_factory

User = get_user_model()

import os
from io import BytesIO
from django.conf import settings
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Image
from django.core.mail import EmailMessage
from .models import MaterialRelease
from django.contrib.auth.decorators import user_passes_test
from django.db import transaction
from accounts.utils import send_kakao_message
from django.db.models import Sum, F, Case, When, Value, IntegerField, ExpressionWrapper
from django.contrib.humanize.templatetags.humanize import intcomma
from django.utils import timezone
from .forms import MaterialReleaseEstimateForm, MaterialReleaseItemIncludeForm

from .utils_pdf import (_ensure_korean_fonts,number_to_korean_amount,_vendor_info,_make_estimate_title,)
import re
from itertools import groupby
from xml.sax.saxutils import escape
from .models import Material, MaterialHistory, MaterialOrderItem
from .utils import redirect_with_filters, log_material_history
from urllib.parse import urlencode
from copy import deepcopy
from reportlab.platypus import PageBreak, Paragraph, Table


def extract_group_name(material_name: str, vendor_type_name: str) -> str:
    """
    교구재명에서 그룹명을 추출
    - 거래처 종류(vendor_type_name)가 '로봇'이면 → 교구재명 첫 번째 단어만 그룹명
    - 그 외의 거래처 종류면 → 전체 교구재명을 그룹명으로 사용
    """
    name = (material_name or "").strip()
    if not name:
        return ""

    if vendor_type_name == "로봇":
        return name.split()[0]  # 첫 번째 단어만
    return name  # 전체 교구명 반환

# ✅ 관리자만 접근 가능하게 하는 함수
def _get_release_admin_user():
    admin_user = User.objects.filter(username="robotmaking@naver.com").first()
    if admin_user:
        return admin_user
    return User.objects.filter(is_staff=True).order_by("id").first()


def _build_release_admin_message(release, teacher_user, item_summary_map, total_quantity):
    created_at = timezone.localtime(release.created_at)
    teacher_name = teacher_user.first_name or teacher_user.get_full_name() or teacher_user.username
    summary_lines = [
        f"- {name} {qty}개"
        for name, qty in sorted(item_summary_map.items())
    ]
    items_text = "\n".join(summary_lines) if summary_lines else "- 품목 없음"

    return (
        "[출고 등록 알림]\n"
        f"주문일시: {created_at.strftime('%Y-%m-%d %H:%M')}\n"
        f"등록자: {teacher_name}\n"
        f"출고처: {release.institution.name}\n"
        f"주문월: {release.order_month}\n"
        f"출고예정일: {str(release.expected_date) if release.expected_date else '-'}\n"
        f"총 수량: {total_quantity}개\n"
        "------------------\n"
        f"{items_text}"
    )


def is_admin(user):
    return user.is_staff or user.is_superuser

@login_required
def toggle_payment_item(request, item_id):
    item = get_object_or_404(MaterialOrderItem, id=item_id)
    if item.paid_date:
        # 이미 수금된 경우 → 다시 미수금으로 초기화
        item.paid_date = None
        messages.info(request, "입금 상태가 '미수금'으로 변경되었습니다.")
    else:
        # 수금 처리 → 오늘 날짜 기록
        item.paid_date = timezone.now().date()
        messages.success(request, "입금이 완료되었습니다.")
    item.save()
    


from django.urls import reverse
from django.contrib.auth.decorators import login_required
from urllib.parse import urlencode

@login_required
def _legacy_create_release_old(request):
    
    selected_teacher_id = request.GET.get("teacher")
    selected_institution_id = request.GET.get("institution")
    
    vendor_types = VendorType.objects.all()
    vendors = Vendor.objects.select_related('vendor_type').all()
    materials = Material.objects.all().order_by("vendor", "vendor_order", "name")

    vendor_kinds = Vendor.objects.exclude(vendor_type__isnull=True) \
        .values_list('vendor_type__name', flat=True).distinct()

    if request.user.is_staff:
        teachers = User.objects.filter(profile__user_type='teacher').order_by('first_name')
        center_teachers = User.objects.filter(profile__user_type='center_teacher').order_by('first_name')
        institutions = TeachingInstitution.objects.all()
    else:
        teachers = User.objects.filter(id=request.user.id, profile__user_type='teacher')
        center_teachers = User.objects.filter(id=request.user.id, profile__user_type='center_teacher')
        institutions = TeachingInstitution.objects.filter(teacher=request.user)

    # 기본값
    row_count = int(request.POST.get("row_count", 1))
    rows_data = []
    for i in range(1, row_count + 1):
        rows_data.append({
            "index": i,
            "vendor": request.POST.get(f"vendor_{i}", ""),
            "material": request.POST.get(f"material_{i}", ""),
            "unit_price": request.POST.get(f"unit_price_{i}", ""),
            "quantity": request.POST.get(f"quantity_{i}", ""),
            "release_method": request.POST.get(f"release_method_{i}", "택배"),
        })

    base_context = {
        'vendor_types': vendor_types,
        'vendor_kinds': vendor_kinds,
        'vendors': vendors,
        'materials': materials,
        'teachers': teachers,
        'center_teachers': center_teachers,
        'institutions': institutions,
        'form_data': request.POST if request.method == "POST" else {},
        'row_count': row_count,
        'rows_data': rows_data,
        "selected_teacher_id": selected_teacher_id,
        "selected_institution_id": selected_institution_id,
    }

    if request.method == 'POST':
        teacher_user = request.user
        if request.user.is_staff:
            teacher_id = request.POST.get('teacher')
            if teacher_id:
                teacher_user = User.objects.get(id=teacher_id)

        institution_id = request.POST.get('institution')
        if not institution_id:
            return render(request, 'release/release_form.html', {
                **base_context,
                'error': '출강 장소를 선택해주세요.',
            })

        institution = TeachingInstitution.objects.get(id=institution_id)

        order_year = request.POST.get('order_year')
        order_month = request.POST.get('order_month')
        if not order_year or not order_month:
            return render(request, 'release/release_form.html', {
                **base_context,
                'error': '올바른 주문 년월을 선택해주세요.',
            })

        order_month = order_month.zfill(2)
        order_month_str = f"{order_year}-{order_month}"

        expected_date = request.POST.get('expected_date')
        if not expected_date:
            return render(request, 'release/release_form.html', {
                **base_context,
                'error': '출고 예정일을 선택해주세요.',
            })

        # 출고 헤더 생성
        release = MaterialRelease.objects.create(
            teacher=teacher_user,
            institution=institution,
            order_month=order_month_str,
            expected_date=expected_date,
            created_by=request.user,   # ✅ 작성자 추가
        )

        any_saved = False
        item_summary = []

        for i in range(1, row_count + 1):
            material_id = request.POST.get(f'material_{i}')
            quantity = request.POST.get(f'quantity_{i}')
            unit_price = request.POST.get(f'unit_price_{i}') or 0
            item_release_method = request.POST.get(f'release_method_{i}', '택배')

            if material_id and quantity and int(quantity) > 0:
                try:
                    material = Material.objects.get(id=int(material_id))
                    group_name = extract_group_name(
                        material.name,
                        material.vendor.vendor_type.name if material.vendor and material.vendor.vendor_type else ""
                    )
                    MaterialReleaseItem.objects.create(
                        release=release,
                        vendor=material.vendor,
                        material=material,
                        unit_price=int(unit_price),
                        quantity=int(quantity),
                        release_method=item_release_method,
                        group_name=group_name,
                    )
                    any_saved = True
                    item_summary.append(f"- {material.name} {quantity}개")
                except Exception as e:
                    print(f"[❌ ERROR] row {i}: {e}")

        if any_saved:
            from accounts.utils import send_kakao_message
            try:
                admin_user = User.objects.get(username="withjongseok@naver.com")
                items_text = "\n".join(item_summary) if item_summary else "품목 없음"
                message = (
                    f"[출고 등록 알림]\n"
                    f"주문자: {teacher_user.first_name or teacher_user.username}\n"
                    f"출강장소: {institution.name}\n"
                    f"주문년월: {order_month_str}\n"
                    f"------------------\n"
                    f"{items_text}"
                )
                send_kakao_message(admin_user, message, local_test=bool(settings.DEBUG))
            except Exception as e:
                print("[카카오 알림 오류]", e)

            # ✅ 기존 필터 유지 redirect
            query_string = request.GET.urlencode()   # 빈 값도 포함한 전체 쿼리스트링
            url = reverse("release_list")
            if query_string:
                url += f"?{query_string}"
            return redirect(url)

        else:
            release.delete()
            return render(request, 'release/release_form.html', {
                **base_context,
                'error': '출고 항목을 입력해주세요.',
            })

    return render(request, 'release/release_form.html', base_context)



@login_required
def create_release(request):
    selected_teacher_id = request.GET.get("teacher")
    selected_institution_id = request.GET.get("institution")

    vendor_types = VendorType.objects.all()
    vendors = Vendor.objects.select_related("vendor_type").all()
    materials = Material.objects.all().order_by("vendor", "vendor_order", "name")
    vendor_kinds = (
        Vendor.objects.exclude(vendor_type__isnull=True)
        .values_list("vendor_type__name", flat=True)
        .distinct()
    )

    if request.user.is_staff:
        teachers = User.objects.filter(profile__user_type="teacher").order_by("first_name")
        center_teachers = User.objects.filter(profile__user_type="center_teacher").order_by("first_name")
        institutions = TeachingInstitution.objects.all()
    else:
        teachers = User.objects.filter(id=request.user.id, profile__user_type="teacher")
        center_teachers = User.objects.filter(id=request.user.id, profile__user_type="center_teacher")
        institutions = TeachingInstitution.objects.filter(teacher=request.user)

    row_count = int(request.POST.get("row_count", 1))
    rows_data = []
    for i in range(1, row_count + 1):
        rows_data.append(
            {
                "index": i,
                "vendor": request.POST.get(f"vendor_{i}", ""),
                "material": request.POST.get(f"material_{i}", ""),
                "unit_price": request.POST.get(f"unit_price_{i}", ""),
                "quantity": request.POST.get(f"quantity_{i}", ""),
                "release_method": request.POST.get(f"release_method_{i}", "택배"),
            }
        )

    base_context = {
        "vendor_types": vendor_types,
        "vendor_kinds": vendor_kinds,
        "vendors": vendors,
        "materials": materials,
        "teachers": teachers,
        "center_teachers": center_teachers,
        "institutions": institutions,
        "form_data": request.POST if request.method == "POST" else {},
        "row_count": row_count,
        "rows_data": rows_data,
        "selected_teacher_id": selected_teacher_id,
        "selected_institution_id": selected_institution_id,
    }

    if request.method == "POST":
        teacher_user = request.user
        if request.user.is_staff:
            teacher_id = request.POST.get("teacher")
            if teacher_id:
                teacher_user = User.objects.get(id=teacher_id)

        institution_id = request.POST.get("institution")
        if not institution_id:
            return render(
                request,
                "release/release_form.html",
                {**base_context, "error": "출강 장소를 선택해주세요."},
            )

        institution = TeachingInstitution.objects.get(id=institution_id)

        order_year = request.POST.get("order_year")
        order_month = request.POST.get("order_month")
        if not order_year or not order_month:
            return render(
                request,
                "release/release_form.html",
                {**base_context, "error": "올바른 주문 연월을 선택해주세요."},
            )

        order_month = order_month.zfill(2)
        order_month_str = f"{order_year}-{order_month}"

        expected_date = request.POST.get("expected_date")
        if not expected_date:
            return render(
                request,
                "release/release_form.html",
                {**base_context, "error": "출고 예정일을 선택해주세요."},
            )

        release = MaterialRelease.objects.create(
            teacher=teacher_user,
            institution=institution,
            order_month=order_month_str,
            expected_date=expected_date,
            created_by=request.user,
        )

        any_saved = False
        item_summary_map = defaultdict(int)
        total_quantity = 0

        for i in range(1, row_count + 1):
            material_id = request.POST.get(f"material_{i}")
            quantity = request.POST.get(f"quantity_{i}")
            unit_price = request.POST.get(f"unit_price_{i}") or 0
            item_release_method = request.POST.get(f"release_method_{i}", "택배")

            if material_id and quantity and int(quantity) > 0:
                try:
                    material = Material.objects.get(id=int(material_id))
                    group_name = extract_group_name(
                        material.name,
                        material.vendor.vendor_type.name if material.vendor and material.vendor.vendor_type else "",
                    )
                    quantity_value = int(quantity)
                    MaterialReleaseItem.objects.create(
                        release=release,
                        vendor=material.vendor,
                        material=material,
                        unit_price=int(unit_price),
                        quantity=quantity_value,
                        release_method=item_release_method,
                        group_name=group_name,
                    )
                    any_saved = True
                    item_summary_map[material.name] += quantity_value
                    total_quantity += quantity_value
                except Exception as e:
                    print(f"[release create error] row {i}: {e}")

        if not any_saved:
            release.delete()
            return render(
                request,
                "release/release_form.html",
                {**base_context, "error": "출고 품목을 입력해주세요."},
            )

        # 새출고 시에는 알림 발송을 생략합니다.
        messages.success(request, "새 교구출고 등록이 완료되었습니다.")

        query_string = request.GET.urlencode()
        url = reverse("release_list")
        if query_string:
            url += f"?{query_string}"
        return redirect(url)

    return render(request, "release/release_form.html", base_context)


@login_required
def release_list(request):
    # 오픈뱅킹 수동매칭용 JSON API
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" and request.GET.get("format") == "json":
        from django.db.models import Sum, F as Fexp
        unpaid = MaterialRelease.objects.filter(payment_status="unpaid").select_related("institution")
        releases = []
        for r in unpaid:
            total = r.items.aggregate(s=Sum(Fexp('unit_price') * Fexp('quantity')))["s"] or 0
            releases.append({"id": r.id, "institution": r.institution.name if r.institution else "", "order_month": r.order_month, "total": int(total)})
        return JsonResponse({"releases": releases})

    user = request.user
    selected_teacher_id = request.GET.get('teacher')
    selected_order_month = (
        request.GET.get('order_month')
        or request.GET.get('order_month_desktop')
        or request.GET.get('order_month_mobile')
    )
    selected_institution_id = request.GET.get('institution')
    unpaid_filter = request.GET.get('unpaid')

    has_pending_subq = MaterialReleaseItem.objects.filter(
        release=OuterRef('pk'),
        status='waiting'
    )

    releases = (
        MaterialRelease.objects
        .select_related('institution', 'teacher')
        .prefetch_related('items__vendor__vendor_type', 'items__material')
        .annotate(has_pending=Exists(has_pending_subq))
    )
    all_releases = releases

    # Institution list for filters
    institutions = TeachingInstitution.objects.filter(is_closed=False).select_related('teacher', 'school').order_by('school__name', 'name', 'program')
    
    selected_institution = None
    if selected_institution_id and selected_institution_id.isdigit():
        try:
            selected_institution = TeachingInstitution.objects.get(id=int(selected_institution_id))
        except TeachingInstitution.DoesNotExist:
            pass
    # Selected teacher's institutions (for admin view)
    selected_teacher_institutions = None
    if user.is_staff and selected_teacher_id:
        try:
            selected_teacher_institutions = institutions.filter(teacher_id=int(selected_teacher_id))
        except (ValueError, TypeError):
            pass

    # Teacher filter (result filtering)
    if selected_teacher_id:
        try:
            releases = releases.filter(teacher_id=int(selected_teacher_id))
        except (ValueError, TypeError):
            pass

    # Institution filter (result filtering)
    if selected_institution_id and selected_institution_id.isdigit():
        releases = releases.filter(institution_id=int(selected_institution_id))

    # Staff vs Teacher logic
    teachers = None
    if user.is_staff:
        teachers = User.objects.filter(profile__user_type='teacher').order_by('first_name')
        center_teachers = User.objects.filter(profile__user_type='center_teacher').order_by('first_name')
    else:
        # Teacher only sees their own data
        releases = releases.filter(teacher=user)
        institutions = institutions.filter(teacher=user)
        selected_teacher_id = user.id
        center_teachers = User.objects.filter(id=user.id, profile__user_type='center_teacher')

    if selected_institution_id and selected_institution_id.isdigit():
        all_releases = all_releases.filter(institution_id=int(selected_institution_id))

    if selected_teacher_id:
        try:
            all_releases = all_releases.filter(teacher_id=int(selected_teacher_id))
        except (ValueError, TypeError):
            pass

    if not user.is_staff:
        all_releases = all_releases.filter(teacher=user)

    # Order month filter
    if selected_order_month:
        releases = releases.filter(order_month=selected_order_month.strip())
        all_releases = all_releases.filter(order_month=selected_order_month.strip())

    unpaid_group_keys = {
        (order.order_month, order.institution_id)
        for order in all_releases.filter(payment_status="unpaid")
    }
    unpaid_group_count = len(unpaid_group_keys)
        
    # Unpaid filter
    if unpaid_filter == "unpaid":
        releases = releases.filter(payment_status="unpaid")
        all_releases = all_releases.filter(payment_status="unpaid")
    elif unpaid_filter == "paid":
        releases = releases.filter(payment_status="paid")
        all_releases = all_releases.filter(payment_status="paid")

    # Order
    releases = releases.order_by('-order_month', '-has_pending', '-created_at').distinct()

    # Grouping/Sums
    grouped_data = {}
    for order in releases:
        key = f"{order.order_month}_{order.institution_id}"
        if key not in grouped_data:
            inst = order.institution
            program_display = inst.program or ''

            grouped_data[key] = {
                "order_month": order.order_month,
                "institution": inst,
                "teacher": order.teacher,
                "orders": [],
                "total_qty": 0,
                "total_price": 0,
                "total_supply": 0,
                "payment_status": getattr(order, 'payment_status', 'unpaid'),
                "payment_date": getattr(order, 'payment_date', None),
                "estimate_sent": True,
                "tax_invoice_sent": False,
                "program_display": program_display,
                "materials_summary": {},
                "items_detail": {},
                "source_type": getattr(order, 'source_type', ''),
                "request_sent": order.request_sent,
            }
            
        if order.payment_status == "paid" and not order.payment_date:
            grouped_data[key]["payment_status"] = "paid"

        if not bool(getattr(order, 'estimate_sent', False)):
            grouped_data[key]["estimate_sent"] = False
        if bool(getattr(order, 'tax_invoice_sent', False)):
            grouped_data[key]["tax_invoice_sent"] = True
        if getattr(order, 'source_type', '') == 'levelup_auto':
            grouped_data[key]["source_type"] = 'levelup_auto'

        for item in order.items.all():
            grouped_data[key]["total_qty"] += item.quantity
            grouped_data[key]["total_price"] += (item.unit_price or 0) * item.quantity
            grouped_data[key]["total_supply"] += (getattr(item.material, 'supply_price', 0) or 0) * item.quantity

            mat_name = item.material.name
            grouped_data[key]["materials_summary"][mat_name] = grouped_data[key]["materials_summary"].get(mat_name, 0) + item.quantity
            # 세금계산서용: 견적서와 동일하게 group_name 기준, included=True만
            if not getattr(item, 'included', True):
                continue
            unit_price = item.unit_price or 0
            row_name = (getattr(item, 'group_name', None) or '').strip() or mat_name
            row_key = (unit_price, row_name)
            if row_key not in grouped_data[key]["items_detail"]:
                grouped_data[key]["items_detail"][row_key] = {"name": row_name, "qty": 0, "unit_price": unit_price, "subtotal": 0}
            grouped_data[key]["items_detail"][row_key]["qty"] += item.quantity
            grouped_data[key]["items_detail"][row_key]["subtotal"] += unit_price * item.quantity

        grouped_data[key]["orders"].append(order)

    for g in grouped_data.values():
        # items_detail을 견적서와 동일하게 정렬된 리스트로 변환
        sorted_rows = sorted(g["items_detail"].values(), key=lambda r: (r["unit_price"], r["name"]))
        tax_total = sum(r["subtotal"] for r in sorted_rows)
        for r in sorted_rows:
            r["supply_subtotal"] = round(r["subtotal"] / 1.1)
            r["vat_subtotal"] = r["subtotal"] - r["supply_subtotal"]
        g["items_detail"] = sorted_rows  # 리스트로 교체
        # 세금계산서 합계는 견적서 포함 항목 기준
        g["tax_supply_total"] = round(tax_total / 1.1)
        g["tax_vat"] = tax_total - g["tax_supply_total"]
        g["tax_total"] = tax_total
        # 기존 total_price 기준 합계도 유지 (하위 호환)
        g["supply_price_total"] = g["tax_supply_total"]
        g["vat"] = g["tax_vat"]
        g["total_with_vat"] = tax_total

    grouped_list = []
    teacher_unpaid_counts = {}
    institution_unpaid_counts = {}
    unpaid_releases = MaterialRelease.objects.filter(payment_status="unpaid").values_list('teacher_id', 'institution_id')
    
    for teacher_id, institution_id in unpaid_releases:
        if teacher_id:
            teacher_unpaid_counts[teacher_id] = teacher_unpaid_counts.get(teacher_id, 0) + 1
        if institution_id:
            institution_unpaid_counts[institution_id] = institution_unpaid_counts.get(institution_id, 0) + 1

    for data in grouped_data.values():
        data["materials_summary"] = dict(sorted(data["materials_summary"].items()))
        data["is_new"] = data["payment_status"] == "unpaid"
        grouped_list.append(data)

    if user.is_staff:
        teachers = User.objects.filter(profile__user_type='teacher').order_by('first_name')
        center_teachers = User.objects.filter(profile__user_type='center_teacher').order_by('first_name')
    else:
        teachers = User.objects.filter(id=user.id, profile__user_type='teacher')
        center_teachers = User.objects.filter(id=user.id, profile__user_type='center_teacher')

    show_teacher_panel = False
    show_school_panel = False
    show_center_panel = False

    if user.is_staff:
        if not selected_institution_id or selected_teacher_id:
            show_teacher_panel = True
        elif selected_institution and selected_institution.place_type in ['school', 'kindergarten']:
            show_school_panel = True
        elif selected_institution and selected_institution.place_type in ['child_center', 'culture_center', 'other']:
            show_center_panel = True
    else:
        if not selected_institution_id or (selected_institution and selected_institution.place_type in ['school', 'kindergarten']):
            show_school_panel = True
        elif selected_institution and selected_institution.place_type in ['child_center', 'culture_center', 'other']:
            show_center_panel = True

    return render(request, 'release/release_list.html', {
        'grouped_list': grouped_list,
        'teachers': teachers,
        'center_teachers': center_teachers,
        'institutions': institutions,
        'selected_teacher_institutions': selected_teacher_institutions,
        'selected_teacher_id': selected_teacher_id,
        'selected_order_month': selected_order_month,
        'selected_institution_id': selected_institution_id,
        'selected_institution': selected_institution,
        'unpaid_filter': unpaid_filter,
        'teacher_unpaid_counts': teacher_unpaid_counts,
        'institution_unpaid_counts': institution_unpaid_counts,
        'unpaid_group_count': unpaid_group_count,
        'show_teacher_panel': show_teacher_panel,
        'show_school_panel': show_school_panel,
        'show_center_panel': show_center_panel,
    })
    
    
@login_required
@user_passes_test(is_admin)
@require_POST
def toggle_release_method(request, item_id):
    item = get_object_or_404(MaterialReleaseItem.objects.select_related("material", "release"), id=item_id)
    material = item.material
    old_method = item.release_method
    new_method = '센터수령' if old_method == '택배' else '택배'

    if item.status == 'released' and material and hasattr(material, 'stock'):
        if old_method == '택배' and new_method == '센터수령':
            if (material.stock or 0) < (item.quantity or 0):
                messages.error(request, f"재고가 부족합니다. (현재 재고: {material.stock})")
                return redirect_with_filters(request, "release_list")

            old_stock = material.stock
            material.stock = F('stock') - (item.quantity or 0)
            material.save(update_fields=['stock'])
            material.refresh_from_db(fields=["stock"])
            log_material_history(
                material=material,
                user=request.user,
                change_type="stock_decrease",
                old_value=old_stock,
                new_value=material.stock,
                note=f"출고방법 변경(택배→센터수령, 출고항목ID {item.id}, 수량 {item.quantity})",
            )
        elif old_method == '센터수령' and new_method == '택배':
            old_stock = material.stock
            material.stock = F('stock') + (item.quantity or 0)
            material.save(update_fields=['stock'])
            material.refresh_from_db(fields=["stock"])
            log_material_history(
                material=material,
                user=request.user,
                change_type="stock_restore",
                old_value=old_stock,
                new_value=material.stock,
                note=f"출고방법 변경(센터수령→택배, 출고항목ID {item.id}, 수량 {item.quantity})",
            )

    item.release_method = new_method
    item.save(update_fields=["release_method"])
    messages.success(request, f"출고방법이 {new_method}로 변경되었습니다.")
    return redirect_with_filters(request, "release_list")


from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponseNotAllowed

@login_required
def set_payment_date(request, institution_id, order_month):
    if request.method == "POST":
        date_val = request.POST.get("payment_date", "").strip()
        action = request.POST.get("action", "")  # ✅ 버튼 구분: 'mark_paid' or 'mark_unpaid'

        releases = MaterialRelease.objects.filter(
            institution_id=institution_id,
            order_month=order_month
        )

        # ✅ '미수금' 버튼 클릭 시 날짜 없더라도 수금완료로 처리
        if action == "mark_paid":
            releases.update(payment_date=None, payment_status="paid")
        elif action == "mark_unpaid":
            releases.update(payment_date=None, payment_status="unpaid")
        elif date_val:
            # ✅ 날짜 입력 시 수금완료
            releases.update(payment_date=date_val, payment_status="paid")
        else:
            # ✅ 날짜 없고 버튼도 없으면 미수금
            releases.update(payment_date=None, payment_status="unpaid")

    # ✅ 돌아갈 URL: Referer 우선 → 없으면 release_list
    referer = request.META.get("HTTP_REFERER")
    if referer:
        return redirect(referer)
    return redirect("release_list")

@login_required
def request_release_notification(request, institution_id, order_month):
    """출고 알림 발송: 강사→관리자(출고요청), 관리자→강사(출고안내)"""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    from django.conf import settings

    # 관리자 계정
    admin_user = User.objects.filter(username="robotmaking@naver.com").first()
    if not admin_user:
        admin_user = User.objects.filter(is_staff=True).first()
    if not admin_user:
        messages.error(request, "관리자 계정을 찾을 수 없습니다.")
        return redirect("release_list")

    if request.user.is_staff:
        # ── 관리자가 강사에게 출고안내 발송 ──────────────────────────
        releases = MaterialRelease.objects.filter(
            institution_id=institution_id,
            order_month=order_month,
        )
        if not releases.exists():
            messages.error(request, "해당하는 출고 내역이 없습니다.")
            return redirect("release_list")

        # 해당 그룹의 담당 강사
        teacher_user = releases.first().teacher
        if not teacher_user:
            messages.error(request, "강사 정보를 찾을 수 없습니다.")
            return redirect("release_list")

        institution = releases.first().institution
        mat_counts = {}
        for r in releases:
            for item in r.items.all():
                mat_name = item.material.name
                mat_counts[mat_name] = mat_counts.get(mat_name, 0) + item.quantity

        items_text = "\n".join(f"- {name} {qty}개" for name, qty in sorted(mat_counts.items()))
        message = (
            f"[출고 안내]\n"
            f"강사: {teacher_user.first_name or teacher_user.username}\n"
            f"출강장소: {institution.name}\n"
            f"주문년월: {order_month}\n"
            f"------------------\n"
            f"{items_text}\n\n"
            f"교구재 출고 준비가 완료되었습니다."
        )

        teacher_profile = getattr(teacher_user, "profile", None)
        if not teacher_profile or not teacher_profile.kakao_id:
            messages.error(request, "강사의 카카오 계정이 연동되어 있지 않습니다.")
            return redirect("release_list")

        from accounts.utils import find_friend_uuid, send_kakao_friend_message
        receiver_uuid = find_friend_uuid(admin_user, teacher_profile.kakao_id)

        if not receiver_uuid:
            messages.error(request, "강사가 관리자의 친구 목록에 없습니다. 카카오톡 친구 추가 및 메시지 권한 동의가 필요합니다.")
            return redirect("release_list")

        res = send_kakao_friend_message(admin_user, receiver_uuid, message, local_test=bool(settings.DEBUG))

        if res and res.get("error"):
            messages.error(request, f"카카오 알림 발송 중 오류가 발생했습니다: {res.get('error')}")
        else:
            messages.success(request, f"{teacher_user.first_name or teacher_user.username} 강사에게 출고 안내를 보냈습니다.")

    else:
        # ── 강사가 관리자에게 출고요청 발송 ──────────────────────────
        releases = MaterialRelease.objects.filter(
            institution_id=institution_id,
            order_month=order_month,
            teacher=request.user,
        )
        if not releases.exists():
            messages.error(request, "해당하는 출고 내역이 없습니다.")
            return redirect("release_list")

        institution = releases.first().institution
        mat_counts = {}
        for r in releases:
            for item in r.items.all():
                mat_name = item.material.name
                mat_counts[mat_name] = mat_counts.get(mat_name, 0) + item.quantity

        items_text = "\n".join(f"- {name} {qty}개" for name, qty in sorted(mat_counts.items()))
        is_re_request = releases.first().request_sent
        message_title = "[출고 재요청 알림]" if is_re_request else "[출고 요청 알림]"
        message = (
            f"{message_title}\n"
            f"강사: {request.user.first_name or request.user.username}\n"
            f"출강장소: {institution.name}\n"
            f"주문년월: {order_month}\n"
            f"------------------\n"
            f"{items_text}"
        )

        admin_profile = getattr(admin_user, "profile", None)
        if not admin_profile or not admin_profile.kakao_id:
            messages.error(request, "관리자의 카카오 계정이 연동되어 있지 않습니다.")
            return redirect("release_list")

        from accounts.utils import find_friend_uuid, send_kakao_friend_message
        receiver_uuid = find_friend_uuid(request.user, admin_profile.kakao_id)

        if not receiver_uuid:
            messages.error(request, "관리자가 친구 목록에 없거나 친구 메시지 권한이 없습니다.")
            return redirect("release_list")

        res = send_kakao_friend_message(request.user, receiver_uuid, message, local_test=bool(settings.DEBUG))

        if res and res.get("error"):
            messages.error(request, f"카카오 알림 발송 중 오류가 발생했습니다: {res.get('error')}")
        else:
            releases.update(request_sent=True)
            messages.success(request, "관리자에게 출고 요청 알림을 보냈습니다.")

    referer = request.META.get("HTTP_REFERER")
    if referer:
        return redirect(referer)
    return redirect("release_list")


@login_required
def order_list(request):
    vendor_id_filter = request.GET.get('vendor_id', '')

    # UI 상단 필터를 위한 데이터
    from .models import Vendor
    vendors = Vendor.objects.all()

    # 모든 주문 아이템 조회 (날짜 내림차순 정렬)
    items_query = (
        MaterialOrderItem.objects
        .filter(order__receive_type='order')
        .select_related('order', 'order__teacher', 'vendor', 'vendor__vendor_type', 'material')
        .order_by('-order__ordered_date', 'vendor__name', 'material__name')
    )

    if vendor_id_filter:
        items_query = items_query.filter(vendor_id=vendor_id_filter)

    from itertools import groupby
    raw_items = list(items_query)
    
    table_rows = []
    total_qty = 0
    total_sum = 0
    
    # 주문일자와 강사(주문자)를 기준으로 함께 그룹화
    for (date_obj, teacher_obj), group in groupby(raw_items, key=lambda i: (i.order.ordered_date, i.order.teacher)):
        group_items = list(group)
        sub_qty = sum(i.quantity for i in group_items if i.quantity)
        sub_sum = sum((i.material.supply_price or 0) * (i.quantity or 0) for i in group_items)
        
        for item in group_items:
            notes = item.order.notes or ""
            auto_suffix = " 출고에 의한 자동생성"
            
            if item.order.release_location:
                item.release_location = item.order.release_location
                item.display_notes = notes
            elif notes.endswith(auto_suffix):
                item.release_location = notes[:-len(auto_suffix)]
                item.display_notes = ""
            else:
                item.release_location = ""
                item.display_notes = notes
                
            table_rows.append({"type": "item", "data": item})
            
        table_rows.append({
            "type": "subtotal",
            "date": date_obj,
            "teacher": teacher_obj,
            "qty": sub_qty,
            "sum": sub_sum
        })
        
        total_qty += sub_qty
        total_sum += sub_sum

    context = {
        "table_rows": table_rows,
        "total_qty": total_qty,
        "total_sum": total_sum,
        "vendors": vendors,
        "selected_vendor_id": vendor_id_filter,
    }
    return render(request, "order/order_list.html", context)

@login_required
def change_payment_status(request, release_id, status):
    release = get_object_or_404(MaterialRelease, id=release_id)
    if status in dict(MaterialRelease.PAYMENT_STATUS_CHOICES):
        release.payment_status = status
        release.save()
    return redirect('release_list')

@login_required
def generate_estimate(request, institution_id, order_month):
    # 1) 폰트
    base_font, base_font_bold = _ensure_korean_fonts()

    # 2) 데이터 수집 (주문년월 + 기관)
    releases = (MaterialRelease.objects
                .filter(institution_id=institution_id, order_month=order_month)
                .prefetch_related('items__vendor__vendor_type', 'items__material', 'institution', 'teacher')
                .order_by('created_at'))
    if not releases.exists():
        return redirect('release_list')

    institution = releases.first().institution
    teacher = releases.first().teacher
    doc_no = f"M{order_month.replace('-', '')}"  # 예: M202508
    settle_date = order_month.replace('-', '.')+"."
    vendor = _vendor_info(institution)   

    # 품목 테이블 데이터
    data = [['NO.', '품 명', '단위', '수량', '단가', '공급가액', '세액', '소계']]
    total_qty = 0
    supply_sum = 0

    idx = 1
    for order in releases:
        for item in order.items.all():
            unit = int(item.unit_price or 0)
            qty = int(item.quantity or 0)
            supply = int(round(unit * qty / 1.1))
            vat = (unit * qty) - supply 
            subtotal = supply + vat                  # 소계 (품목별 합계)
            
            data.append([
                str(idx),
                getattr(item.material, 'name', ''),
                'EA',
                f"{qty}",
                f"{unit:,}",
                f"{supply:,}",
                f"{vat:,}",
                f"{subtotal:,}"
            ])
            idx += 1
            total_qty += qty
            supply_sum += supply
            vat_sum += vat 

    grand_total = supply_sum + vat_sum
    hangul_total = number_to_korean_amount(grand_total)

    # 3) PDF 문서 & 스타일
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm
    )
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('KTitle', parent=styles['Title'], fontName=base_font_bold, fontSize=20, leading=24)
    style_label = ParagraphStyle('KLabel', parent=styles['Normal'], fontName=base_font_bold, fontSize=10)
    style_text = ParagraphStyle('KText', parent=styles['Normal'], fontName=base_font, fontSize=10)
    style_small = ParagraphStyle('KSmall', parent=styles['Normal'], fontName=base_font, fontSize=9)

    story = []

    # ---------- 상단 문서번호/제목/결제일자 ----------
    # ---------- 상단 문서번호(좌) / 견적서(정중앙) ----------
    total_w = doc.width                 # 사용할 가로 폭 (여백 제외)
    side_w  = total_w * 0.25            # 좌/우 동일 폭
    mid_w   = total_w - (side_w * 2)    # 가운데 폭

    top_table = Table([[
        Paragraph(f"No. 문서번호 : {doc_no}", style_small),  # 좌: 문서번호 (왼쪽 정렬)
        Paragraph("견  적  서", style_title),                # 중: 제목 (가운데 정렬)
        Paragraph("", style_small),                          # 우: 비워서 좌우 균형 유지
    ]], colWidths=[side_w, mid_w, side_w])

    top_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'LEFT'),     # 문서번호 왼쪽
        ('ALIGN', (1,0), (1,0), 'CENTER'),   # 제목 중앙
        ('ALIGN', (2,0), (2,0), 'RIGHT'),    # 우측은 비워두되 정렬 지정
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(top_table)
    story.append(Spacer(1, 4))

    # ---------- 수신 / 공급자 박스 ----------
    # 좌측: '귀하' 문구, 우측: 공급자 정보 표
    
    # 1) 학교/기관명 (releases.first()로 이미 institution 확보해 둔 상태)
    school_name = institution.name if institution else ""
    
    # 2) 학교명/귀하 전용 스타일 (학교명을 귀하보다 더 크게)
    style_school = ParagraphStyle(
        name="SchoolName",
        parent=styles['Normal'],
        fontName=base_font_bold,  # 굵게
        fontSize=14,              # 더 크게
        leading=18,
        alignment=TA_LEFT,
    )
    style_to = ParagraphStyle(
        name="ToName",
        parent=styles['Normal'],
        fontName=base_font,
        fontSize=12,              # 학교명보다 작게
        leading=14,
        alignment=TA_LEFT,
    )

    # 3) 수신 박스: [학교명], [귀하], [인사 문구]
    
    style_center = ParagraphStyle(
    name="KCenter",
        parent=styles['Normal'],
        fontName=base_font,
        fontSize=10,
        alignment=TA_CENTER
    )
    style_center_bold = ParagraphStyle(
        name="KCenterBold",
        parent=styles['Normal'],
        fontName=base_font_bold,
        fontSize=14,
        alignment=TA_CENTER
    )
    
    receiver = Table([
        [Paragraph(f"견적일자 : {settle_date}", style_center)],
        [Paragraph(f"<font size=14>{school_name}</font> <font size=10>귀하</font>", style_center_bold)],
        [Paragraph('귀 교의 무궁한 발전을 기원합니다.<br/>의뢰하신 견적은 아래와 같습니다.', style_text)],
    ], colWidths=[60*mm])
    
    receiver.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),    # 가로 중앙
        ('TOPPADDING', (0,0), (-1,-1), 16),    # ↑ 기존 6 → 10
        ('BOTTOMPADDING', (0,0), (-1,-1), 15), # ↑ 기존 6 → 10
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('FONTNAME', (0,0), (-1,-1), base_font),
    ]))
    
    # 공급자 로고(도장) 준비
    logo_elem = None
    if vendor.get("logo_path") and os.path.exists(vendor["logo_path"]):
        logo_elem = Image(vendor["logo_path"], width=25*mm, height=25*mm, kind='proportional')

    right_rows = [
        [Paragraph("공<br/><br/>급<br/><br/>자", style_label),
        Paragraph('사업자 등록번호', style_small), Paragraph(vendor["business_no"], style_text), ''],
        ['', Paragraph('상호', style_small), Paragraph(vendor["name"], style_text),
        Paragraph(f"성명: {vendor['ceo']}", style_text)],
        ['', Paragraph('사업장 주소', style_small), Paragraph(vendor["address"], style_text), logo_elem or ''],
        ['', Paragraph('업태', style_small), Paragraph(vendor["biz_type"], style_text), ''],
        ['', Paragraph('전화', style_small), Paragraph(vendor["tel"] + " / " + vendor["mobile"], style_text), ''],
        ['', Paragraph('팩스', style_small), Paragraph(vendor["fax"], style_text), ''],
        ['', Paragraph('입금계좌', style_small), Paragraph(vendor["bank"], style_text), ''],
    ]

    right_table = Table(right_rows, colWidths=[15*mm, 25*mm, 55*mm, 30*mm])
    right_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('GRID', (1,0), (2,-1), 0.25, colors.grey),  # 가운데 2칸만 세로줄
        ('SPAN', (0,0), (0,-1)),   # "공급자" 세로 병합
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('FONTNAME', (0,0), (-1,-1), base_font),
        
        # 제목 셀 배경색
        ('BACKGROUND', (1,0), (1,0), colors.lightgrey),  # 사업자등록번호
        ('BACKGROUND', (1,1), (1,1), colors.lightgrey),  # 상호
        ('BACKGROUND', (1,2), (1,2), colors.lightgrey),  # 사업장주소
        ('BACKGROUND', (1,3), (1,3), colors.lightgrey),  # 업태
        ('BACKGROUND', (1,4), (1,4), colors.lightgrey),  # 전화
        ('BACKGROUND', (1,5), (1,5), colors.lightgrey),  # 팩스
        ('BACKGROUND', (1,6), (1,6), colors.lightgrey),  # 입금계좌
    ]))

    # 로고가 있다면 상단에 작게 넣기(선택)
    if vendor.get("logo_path") and os.path.exists(vendor["logo_path"]):
        logo = Image(vendor["logo_path"], width=20*mm, height=20*mm, kind='proportional')
        # 로고는 공급자 표 위쪽에 별도로 배치할 수도 있음(단순화: 생략 가능)
    # 좌/우 나란히
    top2 = Table([[receiver, right_table]], colWidths=[65*mm, 112*mm])
    top2.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(top2)
    story.append(Spacer(1, 6))

    # ---------- 합계금액 박스 (엑셀의 큰 라인) ----------
    
    # 중앙정렬 스타일 추가
    style_label_center = ParagraphStyle(
        'KLabelCenter',
        parent=style_label,
        alignment=TA_CENTER
    )
    style_text_center = ParagraphStyle(
        'KTextCenter',
        parent=style_text,
        alignment=TA_CENTER
    )

    sum_line = Table([[
        Paragraph('합계금액 (공급가액+세액)', style_label_center),
        Paragraph(f"{hangul_total}  ({grand_total:,}원)", style_text_center),
    ]], colWidths=[80*mm, 92*mm])

    sum_line.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('BACKGROUND', (1,0), (1,0), colors.white),       # 두 번째 셀 배경 흰색
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),  # 세로 가운데
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('FONTNAME', (0,0), (-1,-1), base_font),
    ]))
    story.append(sum_line)
    story.append(Spacer(1, 6))

    # ---------- 품목 테이블 ----------
    tbl = Table(
        data,
        colWidths=[10*mm, 42*mm, 12*mm, 12*mm, 22*mm, 26*mm, 22*mm, 26*mm]
    )
    tbl.setStyle(TableStyle([
        # 헤더
        ('FONTNAME', (0,0), (-1,0), base_font_bold),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,0), 'MIDDLE'),

        # 본문 폰트
        ('FONTNAME', (0,1), (-1,-1), base_font),
        ('FONTSIZE', (0,1), (-1,-1), 10),

        # 정렬
        ('ALIGN', (0,1), (0,-1), 'CENTER'),  # NO.
        ('ALIGN', (1,1), (1,-1), 'CENTER'),  # 품명 ← 가운데 정렬
        ('ALIGN', (2,1), (3,-1), 'CENTER'),  # 단위, 수량
        ('ALIGN', (4,1), (-1,-1), 'RIGHT'),  # 단가~세액
        
        # ✅ 행 높이 키우기
        ('TOPPADDING', (0,0), (-1,-1), 6),     # 기존보다 키움
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),

        # 그리드
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
    ]))
    story.append(tbl)

    # ---------- 합계 행(공급가/세액/총액) ----------
    sum_rows = [
        ['합 계', '', '', f'{total_qty:,}', '', f'{supply_sum:,}', f'{vat_sum:,}', f'{grand_total:,}']
    ]
    sums = Table(sum_rows, colWidths=[10*mm, 42*mm, 12*mm, 12*mm, 22*mm, 26*mm, 22*mm, 26*mm])
    sums.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), base_font),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,0), 'CENTER'),     # "합계" 가운데
        ('ALIGN', (3,0), (3,0), 'CENTER'),     # 수량 가운데
        ('ALIGN', (4,0), (-1,0), 'RIGHT'),     # 단가~소계 모두 오른쪽 정렬
        
        # ✅ 행 높이 키우기
        ('TOPPADDING', (0,0), (-1,-1), 6),     # 기존보다 키움
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(sums)
    
     # 3) PDF 문서 & 스타일
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm
    )
    # ... (스타일/스토리 구성, 테이블들 추가)
    # 견적서 + 납품서 모두 포함
    from reportlab.platypus import PageBreak

    # 복제하여 두 번째 페이지 생성
    delivery_story = []
    for flowable in story:
        if isinstance(flowable, Paragraph):
            new_text = flowable.text.replace("견  적  서", "납  품  서")
            delivery_story.append(Paragraph(new_text, flowable.style))
        else:
            delivery_story.append(flowable)
    delivery_story.insert(0, PageBreak())

    # 최종 PDF 빌드 (두 페이지)
    doc.build(story + delivery_story)
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()
    

    # ✅ 메일 첨부용으로 바이트 꺼내기
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()
    filename = f"{order_month} {institution.name}_{institution.program}_견적서.pdf"

    # ✅ 메일 수신자 구성 (빈 값은 자동 제외)
    recipients = []
    # TeachingInstitution 모델에 contact_email, admin_email 필드가 있다고 가정
    if getattr(institution, "contact_email", None):
        recipients.append(institution.contact_email)
    if getattr(institution, "admin_email", None):
        recipients.append(institution.admin_email)

    # 옵션: 강사도 참조(참고용)
    cc_list = []
    if getattr(teacher, "email", None):
        cc_list.append(teacher.email)

    # ✅ 메일 발송 (수신자가 있을 때만)
    if recipients:
        subject = f"{order_month} {institution.name}_{institution.program}_견적서.pdf"
        body_lines = [
            f"{institution.name} 귀하,",
            "",
            "안녕하세요. 요청하신 견적서를 첨부드립니다.",
            "",
            f"- 주문년월 : {order_month}",
            f"- 공급가액 : {supply_sum:,}원",
            f"- 세    액 : {vat_sum:,}원",
            f"- 합   계 : {grand_total:,}원",
            "",
            "확인 부탁드립니다. 감사합니다."
        ]
        body = "\n".join(body_lines)

        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
            to=recipients,
            cc=cc_list or None,
        )
        email.attach(filename, pdf_bytes, "application/pdf")
        # 실패 시 예외가 나도록(문제 있으면 콘솔/로깅에서 바로 확인)
        email.send(fail_silently=False)
        
        # ✅ 견적 발송 여부 업데이트
        MaterialRelease.objects.filter(
            institution_id=institution_id,
            order_month=order_month
        ).update(estimate_sent=True)

    # 빌드 & 응답
    doc.build(story)
    buffer.seek(0)
    filename = f"{order_month} {institution.name}_{institution.program}_견적서.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)

@login_required
def edit_material_item(request, item_id):
    item = get_object_or_404(MaterialOrderItem, id=item_id)
    # 🔽 거래처별 정렬
    materials = Material.objects.all().order_by("vendor", "vendor_order", "name")

    if request.method == 'POST':
        material_id = request.POST.get('material')
        quantity = request.POST.get('quantity')

        if material_id and quantity:
            material = Material.objects.get(id=material_id)
            item.material = material
            item.vendor = material.vendor  # ✅ 교구재 바꾸면 거래처도 같이 업데이트
            item.quantity = int(quantity)
            item.save()
            return redirect('order_list')

    return render(request, 'order/edit_item_form.html', {
        'item': item,
        'materials': Material.objects.all(),
        'vendors': Vendor.objects.select_related('vendor_type'),
    })


@login_required
@require_POST
def delete_material_item(request, item_id):
    item = get_object_or_404(MaterialOrderItem, id=item_id)
    order = item.order  # 부모 주문 저장

    item.delete()

    # ✅ 아이템이 하나도 안 남으면 부모 주문도 삭제
    if not order.items.exists():
        order.delete()

    messages.success(request, "주문 항목이 삭제되었습니다.")
    return redirect("order_list")


@login_required
def receive_material_item(request, item_id):
    item = get_object_or_404(MaterialOrderItem, id=item_id)

    if item.status != 'received':
        old_stock = item.material.stock  # ✅ 변경 전 수량
        item.status = 'received'
        item.received_date = timezone.now()

        # ✅ 입고 시 재고 증가
        item.material.stock += item.quantity
        item.material.save()
        item.save()

        log_material_history(
            item.material, request.user,
            "stock_increase", old_stock, item.material.stock,
            f"입고 처리 (주문ID {item.id}, 수량 {item.quantity})"
        )

    return redirect('order_list')

def receive_order(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    if request.method == 'POST':
        form = ReceiptForm(request.POST)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.order = order
            receipt.save()
            return redirect('order_list')
    else:
        form = ReceiptForm()
    return render(request, 'order/receive_form.html', {'form': form, 'order': order})

from django.shortcuts import render, redirect
from .models import Order, Material, VendorType

@login_required
def material_search_api(request):
    """교구재 자동완성 검색 API"""
    q = request.GET.get("q", "").strip()
    qs = Material.objects.select_related("vendor__vendor_type").all()
    if q:
        qs = qs.filter(name__icontains=q)
    results = [
        {
            "id": m.id,
            "name": m.name,
            "stock": m.stock,
            "vendor_id": m.vendor.id if m.vendor else None,
            "vendor_name": m.vendor.name if m.vendor else "",
            "vendor_type_id": m.vendor.vendor_type.id if m.vendor and m.vendor.vendor_type else None,
            "vendor_type_name": m.vendor.vendor_type.name if m.vendor and m.vendor.vendor_type else "",
        }
        for m in qs[:30]
    ]
    return JsonResponse({"results": results})


@login_required
def vendor_search_api(request):
    """거래처 자동완성 검색 API"""
    q = request.GET.get("q", "").strip()
    qs = Vendor.objects.select_related("vendor_type").all()
    if q:
        qs = qs.filter(name__icontains=q)
    results = [
        {
            "id": v.id,
            "name": v.name,
            "vendor_type_id": v.vendor_type.id if v.vendor_type else None,
            "vendor_type_name": v.vendor_type.name if v.vendor_type else "",
        }
        for v in qs[:30]
    ]
    return JsonResponse({"results": results})


@login_required
def create_order(request):
    vendor_types = VendorType.objects.all()
    vendors = Vendor.objects.select_related('vendor_type').all()
    # 🔽 거래처별 정렬
    materials = Material.objects.all().order_by("vendor", "vendor_order", "name")

    vendor_kinds = Vendor.objects.exclude(vendor_type__isnull=True) \
        .values_list('vendor_type__name', flat=True).distinct()

    if request.method == 'POST':
        ordered_date = request.POST.get('ordered_date')
        expected_date = request.POST.get('expected_date')
        receive_type = request.POST.get('receive_type', 'order')
        release_location = request.POST.get('release_location', '메듀랩')
        notes = request.POST.get('notes', '')

        if not ordered_date:
            return render(request, 'order/order_form.html', {
                'error': '주문일자를 선택해주세요.',
                'vendor_types': vendor_types,
                'vendor_kinds': vendor_kinds,
                'vendors': vendors,
                'materials': materials,
                'today': timezone.now().date(),
            })

        if not expected_date:
            return render(request, 'order/order_form.html', {
                'error': '예상 입고일을 선택해주세요.',
                'vendor_types': vendor_types,
                'vendor_kinds': vendor_kinds,
                'vendors': vendors,
                'materials': materials,
                'today': timezone.now().date(),
            })

        # ✅ 동적으로 넘어온 material_X, quantity_X 수집
        items_to_save = []
        for key, value in request.POST.items():
            if key.startswith("material_") and value:
                row_num = key.split("_")[1]
                material_id = value
                quantity = request.POST.get(f"quantity_{row_num}")
                if quantity and quantity.isdigit() and int(quantity) > 0:
                    try:
                        material = Material.objects.get(id=int(material_id))
                        items_to_save.append((material, int(quantity)))
                    except Exception as e:
                        print(f"[❌ ERROR] {row_num}행 처리 실패: {e}")

        if not items_to_save:
            return render(request, 'order/order_form.html', {
                'error': '주문 항목을 입력해주세요.',
                'vendor_types': vendor_types,
                'vendor_kinds': vendor_kinds,
                'vendors': vendors,
                'materials': materials,
                'today': timezone.now().date(),
            })

        # ✅ 주문 생성
        order = MaterialOrder.objects.create(
            teacher=request.user,
            ordered_date=ordered_date,
            expected_date=expected_date,
            receive_type=receive_type,
            release_location=release_location,
            notes=notes,
        )

        # ✅ 아이템 저장
        for material, qty in items_to_save:
            item = MaterialOrderItem.objects.create(
                order=order,
                vendor=material.vendor,
                material=material,
                quantity=qty,
                receive_type=receive_type,  # 주문 단위 입고 종류 반영
                notes=notes,                # 주문 비고 반영
            )
            # ✅ 반납입고일 경우 자동으로 입금완료 처리
            if receive_type == "return":
                item.paid_date = timezone.now().date()
                item.save(update_fields=["paid_date"])

        messages.success(request, "주문이 등록되었습니다.")
        return redirect('order_list')

    # GET 요청 시
    return render(request, 'order/order_form.html', {
        'vendor_types': vendor_types,
        'vendor_kinds': vendor_kinds,
        'vendors': vendors,
        'materials': materials,
        'today': timezone.now().date(),
    })




def vendor_type_list(request):
    types = VendorType.objects.all().order_by('name')  # ✅ 이름 오름차순 정렬
    return render(request, 'materials/vendor_type_list.html', {'types': types})


def vendor_type_create(request):
    if request.method == 'POST':
        form = VendorTypeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('vendor_type_list')
    else:
        form = VendorTypeForm()
    return render(request, 'materials/vendor_type_form.html', {'form': form, 'title': '종류 추가'})

def vendor_type_update(request, pk):
    type_obj = get_object_or_404(VendorType, pk=pk)
    if request.method == 'POST':
        form = VendorTypeForm(request.POST, instance=type_obj)
        if form.is_valid():
            form.save()
            return redirect('vendor_type_list')
    else:
        form = VendorTypeForm(instance=type_obj)
    return render(request, 'materials/vendor_type_form.html', {'form': form, 'title': '종류 수정'})

def vendor_type_delete(request, pk):
    type_obj = get_object_or_404(VendorType, pk=pk)
    if request.method == 'POST':
        type_obj.delete()
        return redirect('vendor_type_list')
    return render(request, 'materials/vendor_type_confirm_delete.html', {'type': type_obj})


def material_list(request):
    from .models import Vendor
    selected_vendor_id = request.GET.get('vendor')
    search_query = request.GET.get('q', '')

    vendors = Vendor.objects.all()

    # 🔽 거래처명 → 교구재 순서(vendor_order) → 교구재명 순서로 정렬
    materials = Material.objects.all().order_by("vendor__name", "vendor_order", "name")

    if selected_vendor_id:
        materials = materials.filter(vendor_id=selected_vendor_id)

    if search_query:
        materials = materials.filter(
            Q(name__icontains=search_query) |
            Q(vendor__name__icontains=search_query)
        )

    return render(request, 'materials/material_list.html', {
        'materials': materials,
        'vendors': vendors,
        'selected_vendor_id': selected_vendor_id,
        'search_query': search_query
    })


@login_required
def material_create(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()

            # ✅ next 파라미터 유지
            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            return redirect('material_list')
    else:
        form = MaterialForm()

    return render(request, 'materials/material_form.html', {
        'form': form,
        'title': '교구재 등록'
    })

from django.shortcuts import render, redirect, get_object_or_404
from .forms import VendorForm
from .models import Vendor

from .models import Vendor, VendorType


def vendor_list(request):
    selected_type_id = request.GET.get('type')  # ✅ GET 파라미터: ?type=1
    vendor_types = VendorType.objects.order_by('name')

    if selected_type_id:
        vendors = Vendor.objects.filter(vendor_type_id=selected_type_id)  # ✅ FIXED
    else:
        vendors = Vendor.objects.all()

    context = {
        'vendors': vendors,
        'vendor_types': vendor_types,
        'selected_type_id': selected_type_id,
    }
    return render(request, 'materials/vendor_list.html', context)



def vendor_create(request):
    if request.method == 'POST':
        form = VendorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('vendor_list')
    else:
        form = VendorForm()
    return render(request, 'materials/vendor_form.html', {'form': form, 'title': '거래처 등록'})

def vendor_edit(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()
            return redirect('vendor_list')
    else:
        form = VendorForm(instance=vendor)
    return render(request, 'materials/vendor_form.html', {'form': form, 'title': '거래처 수정'})

def vendor_update(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        form = VendorForm(request.POST, instance=vendor)
        if form.is_valid():
            form.save()
            return redirect('vendor_list')
    else:
        form = VendorForm(instance=vendor)
    return render(request, 'materials/vendor_form.html', {'form': form, 'title': '거래처 수정'})

def vendor_delete(request, pk):
    vendor = get_object_or_404(Vendor, pk=pk)
    if request.method == 'POST':
        vendor.delete()
        return redirect('vendor_list')
    return render(request, 'materials/vendor_confirm_delete.html', {'vendor': vendor})

import openpyxl
from django.contrib import messages
from .forms import MaterialUploadForm, MaterialForm
from .models import Material, Vendor

# 교구재 대량 등록
def material_bulk_upload(request):
    if request.method == 'POST':
        form = MaterialUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # ✅ 엑셀 헤더: 거래처종류, 거래처명, 교구재 이름, 소비자가, 학교납품가, 기관납품가, 공급가, 재고
            for row in sheet.iter_rows(min_row=2, values_only=True):
                vendor_type_name, vendor_name, name, consumer_price, school_price, institute_price, supply_price, stock = row

                # 필수 데이터 없으면 건너뛰기
                if not vendor_type_name or not vendor_name or not name:
                    continue

                # 값들을 깨끗하게 정리 (앞뒤 공백 제거)
                vendor_type_name = str(vendor_type_name).strip()
                vendor_name = str(vendor_name).strip()
                material_name = str(name).strip() # 변수명 중복 피하기 위해 변경

                # 1) 거래처 종류 찾거나 생성
                vendor_type, _ = VendorType.objects.get_or_create(name=vendor_type_name)

                # 2) 거래처 찾거나 생성 (거래처 종류 연결)
                vendor, _ = Vendor.objects.get_or_create(
                    name=vendor_name,
                    vendor_type=vendor_type
                )

                # ✨ 여기에 이미 등록된 교구재인지 확인하는 코드를 추가해요! ✨
                # vendor, material_name, vendor_type를 기준으로 이미 있는지 확인
                existing_material = Material.objects.filter(
                    vendor=vendor,
                    name=material_name,
                    vendor_type=vendor_type # 이전에 생긴 오류를 해결하기 위해 vendor_type도 함께 확인!
                ).first()

                if existing_material:
                    # 이미 존재하는 교구재라면 건너뛰기
                    print(f"Skipping existing material: {material_name} ({vendor_name})") # 콘솔에 표시 (선택 사항)
                    continue # 다음 엑셀 행으로 이동!

                # 3) 새로운 교구재 생성
                Material.objects.create(
                    vendor=vendor,
                    vendor_type=vendor_type, # 중요! 이전에 수정했던 부분이에요!
                    name=material_name,
                    consumer_price=consumer_price or 0,
                    school_price=school_price or 0,
                    institute_price=institute_price or 0,
                    supply_price=supply_price or 0,
                    stock=stock or 0
                )

            messages.success(request, '교구재가 엑셀에서 모두 등록되었습니다. 중복된 교구재는 건너뛰었습니다.')
            return redirect('material_list')
    else:
        form = MaterialUploadForm()

    return render(request, 'materials/material_upload.html', {'form': form})

import openpyxl
from django.http import HttpResponse

# 교구재 템플릿 다운로드
def download_material_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "교구재 템플릿"

    # 헤더 작성 (거래처 종류 추가)
    ws.append(['거래처종류', '거래처명', '교구재 이름', '소비자가', '학교납품가', '기관납품가', '공급가', '재고'])

    # 예시 데이터 (선택)
    ws.append(['로봇', '예시거래처', '예시교구재', 10000, 9000, 8500, 8000, 50])

    # 엑셀 다운로드 응답
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=material_template.xlsx'
    wb.save(response)
    return response

from django.shortcuts import get_object_or_404
from .models import Material, MaterialHistory
#교구재 수정
@login_required
def material_edit(request, pk):
    material = get_object_or_404(Material, pk=pk)
    old_stock = material.stock  # 수정 전 수량 기록

    if request.method == 'POST':
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            updated = form.save()

            # ✅ 수량 변동 기록
            if old_stock != updated.stock:
                MaterialHistory.objects.create(
                    material=updated,
                    user=request.user,
                    change_type="stock_increase" if updated.stock > old_stock else "stock_decrease",
                    old_value=old_stock,
                    new_value=updated.stock,
                    note="수정 화면에서 직접 변경"
                )

            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)
            return redirect("material_list")
    else:
        form = MaterialForm(instance=material)

    return render(request, "materials/material_form.html", {
        "form": form,
        "title": "교구재 수정"
    })



@login_required
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)

    if request.method == "POST":
        material.delete()

        # ✅ next: POST > GET 확인 후 redirect
        next_url = request.POST.get("next") or request.GET.get("next")
        if next_url:
            return redirect(next_url)
        return redirect("material_list")

    # ✅ GET 요청으로 직접 접근하면 리스트로 돌려보냄
    return redirect("material_list")


@login_required
def material_history(request, pk):
    material = get_object_or_404(Material, pk=pk)
    histories = MaterialHistory.objects.filter(material=material).order_by("-created_at")
    return render(request, "materials/material_history.html", {
        "material": material,
        "histories": histories
    })




def _ensure_korean_fonts():
    """
    사용할 한글 폰트를 등록하고 (정상/볼드) 폰트 이름을 반환합니다.
    우선순위: 시스템 설치(Nanum) > 프로젝트 포함 경로(assets/fonts)
    """
    # 후보 경로
    candidates = [
        # Ubuntu fonts-nanum 설치 경로
        ('NanumGothic', '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
         'NanumGothic-Bold', '/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf'),
        # 프로젝트 포함 경로
        ('NanumGothic', os.path.join(settings.BASE_DIR, 'assets', 'fonts', 'NanumGothic.ttf'),
         'NanumGothic-Bold', os.path.join(settings.BASE_DIR, 'assets', 'fonts', 'NanumGothicBold.ttf')),
    ]

    for normal_name, normal_path, bold_name, bold_path in candidates:
        if os.path.exists(normal_path) and os.path.exists(bold_path):
            if normal_name not in pdfmetrics._fonts:
                pdfmetrics.registerFont(TTFont(normal_name, normal_path))
            if bold_name not in pdfmetrics._fonts:
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
            # 폰트 패밀리 매핑(볼드 스타일 지정 쉽게)
            try:
                from reportlab.pdfbase.pdfmetrics import registerFontFamily
                registerFontFamily(normal_name, normal=normal_name, bold=bold_name, italic=normal_name, boldItalic=bold_name)
            except Exception:
                pass
            return normal_name, bold_name

    # 최후 수단: Helvetica(한글 미지원)로 남지만, 경고 삼아 이름 반환
    return 'Helvetica', 'Helvetica-Bold'

def _build_estimate_pdf_bytes(institution_id: int, order_month: str):
    """PDF 바이트와 파일명, 합계 정보를 반환 (기관 + 주문월 전체 기준, (단가, 그룹명/원래명) 묶음)"""
    # 0) 한글 폰트 로드
    base_font, base_font_bold = _ensure_korean_fonts()

    # 1) 데이터 수집: 기관 + 주문월 전체 Release
    releases = (
        MaterialRelease.objects.filter(institution_id=institution_id, order_month=order_month)
        .prefetch_related(
            'items__vendor__vendor_type',
            'items__material',
            'institution', 'teacher'
        )
        .order_by('created_at')
    )
    if not releases.exists():
        # 이 함수는 뷰가 아니므로 redirect 하지 않음
        return None, None, None

    institution = releases.first().institution
    teacher = releases.first().teacher

    # 1-1) 월 전체 아이템 (included=True만 반영)
    items_qs = MaterialReleaseItem.objects.filter(release__in=releases)
    if hasattr(MaterialReleaseItem, 'included'):
        items_qs = items_qs.filter(included=True)
    items_qs = (
        items_qs.select_related('material', 'vendor', 'release')
                .order_by('vendor__name', 'material__name', 'pk')
    )
    if not items_qs.exists():
        return None, None, None
    
    # 제목 자동 생성
    doc_title = _make_estimate_title(institution, order_month)

    # 문서번호/일자/공급자
    doc_no = f"M{order_month.replace('-', '')}"       # 예: M202509
    settle_date = order_month.replace('-', '.') + "."
    vendor = _vendor_info(institution)

    # 2) 표 구성 및 합계 계산
    #    ✅ 묶음 기준 키 = (unit_price, group_name 또는 material.name)
    grouped = defaultdict(lambda: {"qty": 0, "items": []})
    for item in items_qs:
        unit = int(item.unit_price or 0)
        key_name = item.group_name.strip() if getattr(item, "group_name", None) else item.material.name
        key = (unit, key_name)
        grouped[key]["qty"] += int(item.quantity or 0)
        grouped[key]["items"].append(item)

    # 정렬: 단가 오름차순 → 표시명 사전순
    grouped_keys = sorted(grouped.keys(), key=lambda k: (k[0], k[1]))

    
    # 3) PDF 문서 & 스타일 (레이아웃)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm
    )
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('KTitle', parent=styles['Title'], fontName=base_font_bold, fontSize=20, leading=24)
    style_label = ParagraphStyle('KLabel', parent=styles['Normal'], fontName=base_font_bold, fontSize=10)
    style_text = ParagraphStyle('KText', parent=styles['Normal'], fontName=base_font, fontSize=10)
    style_small = ParagraphStyle('KSmall', parent=styles['Normal'], fontName=base_font, fontSize=9,alignment=TA_CENTER)
    # 한국어/긴 단어도 자연스럽게 줄바꿈되도록
    style_wrap = ParagraphStyle(
        'KWrap',
        parent=styles['Normal'],
        fontName=base_font,     # 한글 폰트
        fontSize=10,
        leading=12,
        wordWrap='CJK'          # ✅ 핵심: CJK 줄바꿈
    )
    story = []

    # ---------- 상단 문서번호/제목 ----------
    total_w = doc.width
    side_w  = total_w * 0.25
    mid_w   = total_w - (side_w * 2)
    top_table = Table(
        [[Paragraph(f"No. 문서번호 : {doc_no}", style_small),
          Paragraph("견  적  서", style_title),
          Paragraph("", style_small)]],
        colWidths=[side_w, mid_w, side_w]
    )
    top_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(top_table)
    story.append(Spacer(1, 4))

    # ---------- 수신 / 공급자 박스 ----------
    school_name = institution.name if institution else ""
    style_center = ParagraphStyle('KCenter', parent=styles['Normal'], fontName=base_font, fontSize=10, alignment=TA_CENTER)
    style_center_bold = ParagraphStyle('KCenterBold', parent=styles['Normal'], fontName=base_font_bold, fontSize=14, alignment=TA_CENTER)

    receiver = Table([
        [Paragraph(f"견적일자 : {settle_date}", style_center)],
        [Paragraph(f"<font size=14>{school_name}</font> <font size=10>귀하</font>", style_center_bold)],
        [Paragraph('귀 교의 무궁한 발전을 기원합니다.<br/>의뢰하신 견적은 아래와 같습니다.', style_text)],
    ], colWidths=[60*mm])
    receiver.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 16),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))
    
    data = [['NO.', '품 명', '단위', '수량', '단가', '공급가액', '세액', '소계']]
    total_qty = 0
    supply_sum = 0
    vat_sum = 0

    idx = 1
    for (unit, display_name) in grouped_keys:
        info = grouped[(unit, display_name)]
        qty = info["qty"]

        line_total = unit * qty                # VAT 포함 소계
        line_vat = (line_total + 5) // 11      # ✅ 세법 기준 반올림
        line_supply = line_total - line_vat

        name_para = Paragraph(escape(display_name), style_wrap)
        
        data.append([
            str(idx),
            name_para,                       # ✅ 그룹명 있으면 그룹명, 없으면 원래 교구재명
            'EA',
            f"{qty}",
            f"{unit:,}",
            f"{line_supply:,}",
            f"{line_vat:,}",
            f"{line_total:,}",
        ])

        idx += 1
        total_qty += qty
        supply_sum += line_supply
        vat_sum += line_vat

    grand_total = supply_sum + vat_sum
    hangul_total = number_to_korean_amount(grand_total)

    # 세로 텍스트용 스타일 (가운데 정렬)
    style_vertical = ParagraphStyle(
        'Vertical',
        parent=style_label,          # 기본 라벨 스타일 기반
        fontName=base_font_bold,     # 볼드 폰트
        fontSize=14,                 # 🔹 글자 키움
        leading=16,                  # 🔹 줄 간격도 키움(세로 배치 시 간격)
        alignment=1,                 # CENTER
    )

    # 중앙정렬 텍스트 스타일 (위에서 이미 정의했다면 재정의 불필요)
    style_text_center = ParagraphStyle(
        'KTextCenter',
        parent=style_text,
        alignment=TA_CENTER
    )
    
    # 도장 이미지 준비
    stamp_elem = None
    if vendor.get("logo_path") and os.path.exists(vendor["logo_path"]):
        stamp_elem = Image(vendor["logo_path"], width=18*mm, height=18*mm, kind="proportional")

    # 성명 + 도장 셀
    if stamp_elem:
        ceo_and_stamp = Table(
            [[Paragraph(vendor.get("ceo", "-"), style_text_center), stamp_elem]],
            colWidths=[25*mm, 18*mm],   # 이름 칸은 가변(None), 도장 칸은 고정
            rowHeights=[12*mm]
        )
        ceo_and_stamp.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('LEFTPADDING', (0,0), (0,0), 30),   # 이름쪽 여백 제거
            ('RIGHTPADDING', (0,0), (0,0), -10),
            ('LEFTPADDING', (1,0), (1,0), 0),   # 도장쪽 여백 제거
            ('RIGHTPADDING', (1,0), (1,0), 0),
        ]))
    else:
        ceo_and_stamp = Paragraph(vendor.get("ceo", "-"), style_text_center)


    # 공급자 정보 행
    right_rows = [
        [Paragraph("사업자 등록번호", style_small),
        Paragraph(vendor.get("business_no", "-"), style_text_center), "", ""],

        # 상호 / 상호명 / 성명 / 대표자+도장
        [Paragraph("상호", style_small),
        Paragraph(vendor.get("name", "-"), style_text_center),
        Paragraph("성명", style_small),
        ceo_and_stamp],

        [Paragraph("사업장 주소", style_small),
        Paragraph(vendor.get("address", "-"), style_text_center), "", ""],

        [Paragraph("업태", style_small),
        Paragraph(vendor.get("biz_type", "-"), style_text_center), "", ""],

        [Paragraph("전화", style_small),
        Paragraph(vendor.get("tel", "-") + " / " + vendor.get("mobile", "-"), style_text_center), "", ""],

        [Paragraph("팩스", style_small),
        Paragraph(vendor.get("fax", "-"), style_text_center), "", ""],

        [Paragraph("입금계좌", style_small),
        Paragraph(vendor.get("bank", "-"), style_text_center), "", ""],
    ]

    # 표 정의
    right_table = Table(
        right_rows, 
        colWidths=[28*mm, 28*mm, 23*mm, 28*mm],
        rowHeights=[7*mm] * len(right_rows)
        )
    right_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,1), (2,1), colors.lightgrey),
        ('SPAN', (1,0), (3,0)),
        ('SPAN', (1,2), (3,2)),
        ('SPAN', (1,3), (3,3)),
        ('SPAN', (1,4), (3,4)),
        ('SPAN', (1,5), (3,5)),
        ('SPAN', (1,6), (3,6)),
    ]))



    top2 = Table([[receiver, right_table]], colWidths=[65*mm, 112*mm])
    top2.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(top2)
    story.append(Spacer(1, 6))

    # ---------- 합계금액 박스 ----------
    style_label_center = ParagraphStyle('KLabelCenter', parent=style_label, alignment=TA_CENTER)
    style_text_center = ParagraphStyle('KTextCenter', parent=style_text, alignment=TA_CENTER)
    sum_line = Table(
        [[Paragraph('합계금액 (공급가액+세액)', style_label_center),
          Paragraph(f"{hangul_total}  ({grand_total:,}원)", style_text_center)]],
        colWidths=[80*mm, 92*mm]
    )
    sum_line.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.8, colors.black),
        ('BACKGROUND', (0,0), (0,0), colors.whitesmoke),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(sum_line)
    story.append(Spacer(1, 6))

    # ---------- 품목 테이블 ----------
    tbl = Table(
        data,
        colWidths=[10*mm, 42*mm, 12*mm, 12*mm, 22*mm, 26*mm, 22*mm, 26*mm]
    )
    tbl.setStyle(TableStyle([
        # 헤더
        ('FONTNAME', (0,0), (-1,0), base_font_bold),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,0), 'MIDDLE'),

        # 본문
        ('FONTNAME', (0,1), (-1,-1), base_font),
        ('FONTSIZE', (0,1), (-1,-1), 10),

        ('ALIGN', (0,1), (0,-1), 'CENTER'),  # NO.
        ('ALIGN', (1,1), (1,-1), 'LEFT'),    # 품명은 좌측정렬
        ('ALIGN', (2,1), (3,-1), 'CENTER'),  # 단위/수량
        ('ALIGN', (4,1), (-1,-1), 'RIGHT'),  # 단가~소계

        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),

        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),
    ]))
    story.append(tbl)

    # ---------- 합계 행 ----------
    sum_rows = [['합 계', '', '', f'{total_qty:,}', '', f'{supply_sum:,}', f'{vat_sum:,}', f'{grand_total:,}']]
    sums = Table(sum_rows, colWidths=[10*mm, 42*mm, 12*mm, 12*mm, 22*mm, 26*mm, 22*mm, 26*mm])
    sums.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), base_font),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,0), 'CENTER'),
        ('ALIGN', (3,0), (3,0), 'CENTER'),
        ('ALIGN', (4,0), (-1,0), 'RIGHT'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(sums)

    # =================================================
    # 🔹 두 번째 페이지 : 납품서
    # =================================================
    from copy import deepcopy

    delivery_story = []

    top_table_delivery = Table(
        [[Paragraph(f"No. 문서번호 : {doc_no}", style_small),
        Paragraph("납  품  서", style_title),
        Paragraph("", style_small)]],
        colWidths=[side_w, mid_w, side_w]
    )
    delivery_story.append(PageBreak())
    delivery_story.append(top_table_delivery)
    delivery_story.append(Spacer(1, 4))

    # ✅ 납품서 수신 박스 (견적서와 동일 스타일 적용)
    receiver_delivery = Table([
        [Paragraph(f"납품일자 : {settle_date}", style_center)],
        [Paragraph(f"<font size=14>{school_name}</font> <font size=10>귀하</font>", style_center_bold)],
        [Paragraph('귀 교의 무궁한 발전을 기원합니다.<br/>납품 내역은 아래와 같습니다.', style_text)],
    ], colWidths=[60*mm])

    receiver_delivery.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 16),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
    ]))

    # 오른쪽 공급자 정보 테이블 복제
    top2_delivery = Table([[receiver_delivery, deepcopy(right_table)]], colWidths=[65*mm, 112*mm])
    top2_delivery.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    delivery_story.append(top2_delivery)
    delivery_story.append(Spacer(1, 6))

    # ✅ 깊은 복사로 동일한 레이아웃 확보
    delivery_story.extend([
        deepcopy(sum_line),
        Spacer(1, 6),
        deepcopy(tbl),
        deepcopy(sums),
    ])

    # ---------- PDF 빌드 ----------
    doc.build(story + delivery_story)
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()

    program_name = getattr(institution, 'program', '') or ''
    filename = f"{order_month} {institution.name}_{program_name}_견적서.pdf".replace('__', '_').strip()

    totals = {
        "total_qty": total_qty,
        "supply_sum": supply_sum,
        "vat_sum": vat_sum,
        "grand_total": grand_total,
        "institution": institution,
        "teacher": teacher,
        "order_month": order_month,
        "hangul_total": hangul_total,
    }
    return pdf_bytes, filename, totals



@login_required
def estimate_preview(request, institution_id, order_month):
    # 대표 release (제목/비고 수정용)
    release = MaterialRelease.objects.filter(
        institution_id=institution_id, order_month=order_month
    ).first()
    if not release:
        messages.error(request, "해당 조건의 견적서 데이터를 찾을 수 없습니다.")
        return redirect('release_list')

    # 기관 + 주문월 전체 아이템 queryset
    month_qs = (
        MaterialReleaseItem.objects.filter(
            release__institution_id=institution_id,
            release__order_month=order_month
        )
        .select_related("material", "release", "vendor")
        .annotate(
            line_total=ExpressionWrapper(
                F("unit_price") * F("quantity"), output_field=IntegerField()
            )
        )
        .order_by("vendor__name", "material__name", "pk")
    )

    ItemFormSet = modelformset_factory(
        MaterialReleaseItem,
        form=MaterialReleaseItemIncludeForm,
        extra=0
    )

    # 자동 제목 (항상 계산해 두고 필요 시 사용)
    auto_title = _make_estimate_title(release.institution, order_month)

    if request.method == "POST":
        form = MaterialReleaseEstimateForm(request.POST, instance=release)
        formset = ItemFormSet(request.POST, queryset=month_qs)

        if form.is_valid() and formset.is_valid():
            # 제목이 비어 제출되면 자동 제목으로 저장
            submitted_title = (form.cleaned_data.get("title") or "").strip()
            if not submitted_title:
                form.instance.title = auto_title

            form.save()
            for item_form in formset.forms:
                item = item_form.instance
                item.included = item_form.cleaned_data.get("included", False)
                item.group_name = (item_form.cleaned_data.get("group_name") or "").strip() or item.material.name
                item.save(update_fields=["included", "group_name"])
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({
                    "success": True,
                    "message": "견적서 내용을 저장했습니다.",
                })
            messages.success(request, "견적서 내용(제목/비고)과 교구재 포함여부/그룹명을 저장했습니다.")
            return redirect("estimate_preview", institution_id=institution_id, order_month=order_month)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({
                "success": False,
                "message": "저장에 실패했습니다. 입력값을 다시 확인해주세요.",
                "errors": {
                    "form": form.errors,
                    "formset": formset.errors,
                },
            }, status=400)
        messages.error(request, "저장에 실패했습니다. 입력값을 다시 확인해주세요.")
    else:
        # GET: 저장된 제목이 없으면 자동 제목을 미리 채워서 폼에 표시
        if not (release.title or "").strip():
            form = MaterialReleaseEstimateForm(instance=release, initial={"title": auto_title})
        else:
            form = MaterialReleaseEstimateForm(instance=release)
        formset = ItemFormSet(queryset=month_qs)

    # 합계 (포함된 교구재만 기준)
    included_qs = month_qs.filter(included=True)
    total_sum = included_qs.aggregate(s=Sum(F("unit_price") * F("quantity")))["s"] or 0
    vat_sum = (total_sum + 5) // 11
    supply_sum = total_sum - vat_sum

    # 단가 + 그룹명/원래명 기준 묶음 미리보기
    group_map = defaultdict(lambda: {"qty": 0, "items": []})
    for it in included_qs:
        unit = int(it.unit_price or 0)
        key_name = it.group_name.strip() if getattr(it, "group_name", None) else it.material.name
        key = (unit, key_name)
        group_map[key]["qty"] += int(it.quantity or 0)
        group_map[key]["items"].append(it)

    grouped_rows = []
    for idx, ((unit, key_name), info) in enumerate(sorted(group_map.items(), key=lambda k: (k[0][0], k[0][1])), start=1):
        qty = info["qty"]
        subtotal = unit * qty
        line_vat = (subtotal + 5) // 11
        line_supply = subtotal - line_vat
        grouped_rows.append({
            "no": idx, "name": key_name, "unit": "EA", "qty": qty,
            "unit_price": unit, "supply": line_supply, "vat": line_vat, "subtotal": subtotal
        })

    # PDF URL/파일명/수신자
    pdf_bytes, filename, totals = _build_estimate_pdf_bytes(institution_id, order_month)
    if not pdf_bytes:
        messages.error(request, "해당 조건의 견적서 데이터를 찾을 수 없습니다.")
        return redirect("release_list")

    inst = totals["institution"]
    recipients = [e for e in [getattr(inst, "contact_email", None), getattr(inst, "admin_email", None)] if e]

    estimate_sent = MaterialRelease.objects.filter(
        institution_id=institution_id,
        order_month=order_month,
        estimate_sent=True,
    ).exists()

    return render(request, "release/estimate_preview.html", {
        "release": release,
        "form": form,
        "formset": formset,
        "filename": filename,
        "pdf_url": f"{request.build_absolute_uri().rstrip('/')}/../pdf/".replace("/preview/", "/pdf/"),
        "send_url": f"{request.build_absolute_uri().rstrip('/')}/../send/".replace("/preview/", "/send/"),
        "supply_sum": supply_sum, "vat_sum": vat_sum, "total_sum": total_sum,
        "totals": totals, "recipients": recipients,
        "grouped_rows": grouped_rows,
        "estimate_sent": estimate_sent,
        "selected_teacher_id": request.GET.get("teacher"),
        "selected_institution_id": request.GET.get("institution"),
        "selected_order_month": request.GET.get("order_month"),
    })

@login_required
def estimate_pdf(request, institution_id, order_month):
    pdf_bytes, filename, totals = _build_estimate_pdf_bytes(institution_id, order_month)
    if not pdf_bytes:
        messages.error(request, "PDF 생성 실패")
        return redirect('release_list')

    # 브라우저에서 바로 열어보기 (다운로드 원하면 as_attachment=True)
    return FileResponse(BytesIO(pdf_bytes), as_attachment=False, filename=filename, content_type='application/pdf')

@login_required
def estimate_send(request, institution_id, order_month):
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if request.method != 'POST':
        return redirect('estimate_preview', institution_id=institution_id, order_month=order_month)

    pdf_bytes, filename, totals = _build_estimate_pdf_bytes(institution_id, order_month)
    if not pdf_bytes:
        if is_ajax:
            return JsonResponse({"success": False, "message": "PDF 생성 실패로 메일을 보낼 수 없습니다."}, status=400)
        messages.error(request, "PDF 생성 실패로 메일을 보낼 수 없습니다.")
        return redirect('release_list')

    inst = totals["institution"]
    teacher = totals["teacher"]

    recipients = []
    if getattr(inst, "contact_email", None):
        recipients.append(inst.contact_email)
    if getattr(inst, "admin_email", None):
        recipients.append(inst.admin_email)

    if not recipients:
        if is_ajax:
            return JsonResponse({"success": False, "message": "수신자 이메일(담당자/행정실)이 없어 메일을 보내지 않았습니다."}, status=400)
        messages.warning(request, "수신자 이메일(담당자/행정실)이 없어 메일을 보내지 않았습니다.")
        return redirect('release_list')

    subject = f"[견적서] {inst.name} - {order_month}"
    body = (
        f"{inst.name} 귀하,\n\n"
        "안녕하세요. 요청하신 견적서를 첨부드립니다.\n\n"
        f"- 주문년월 : {order_month}\n"
        f"- 공급가액 : {totals['supply_sum']:,}원\n"
        f"- 세    액 : {totals['vat_sum']:,}원\n"
        f"- 합   계 : {totals['grand_total']:,}원\n\n"
        "확인 부탁드립니다. 감사합니다."
    )

    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
        to=recipients,
        cc=[teacher.email] if getattr(teacher, "email", None) else None,
    )
    email.attach(filename, pdf_bytes, "application/pdf")

    try:
        email.send(fail_silently=False)

        print("이메일 발송 완료!")

        MaterialRelease.objects.filter(
            institution_id=institution_id,
            order_month=order_month
        ).update(estimate_sent=True)

        # 카카오톡 알림: 관리자→강사
        try:
            admin_user = User.objects.filter(username="robotmaking@naver.com").first() \
                         or User.objects.filter(is_staff=True).first()
            teacher_profile = getattr(teacher, "profile", None)
            if admin_user and teacher_profile and teacher_profile.kakao_id:
                from accounts.utils import find_friend_uuid, send_kakao_friend_message
                receiver_uuid = find_friend_uuid(admin_user, teacher_profile.kakao_id)
                if receiver_uuid:
                    mat_counts = {}
                    releases_qs = MaterialRelease.objects.filter(
                        institution_id=institution_id,
                        order_month=order_month,
                    )
                    for r in releases_qs:
                        for item in r.items.filter(included=True):
                            mat_name = item.material.name
                            mat_counts[mat_name] = mat_counts.get(mat_name, 0) + item.quantity
                    items_text = "\n".join(f"- {name} {qty}개" for name, qty in sorted(mat_counts.items()))
                    kakao_msg = (
                        f"[견적서 발송 안내]\n"
                        f"강사: {teacher.first_name or teacher.username}\n"
                        f"출강장소: {inst.name}\n"
                        f"주문년월: {order_month}\n"
                        f"------------------\n"
                        f"{items_text}\n"
                        f"------------------\n"
                        f"공급가액: {totals['supply_sum']:,}원\n"
                        f"세    액: {totals['vat_sum']:,}원\n"
                        f"합   계: {totals['grand_total']:,}원\n\n"
                        f"견적서가 이메일로 발송되었습니다."
                    )
                    send_kakao_friend_message(admin_user, receiver_uuid, kakao_msg, local_test=bool(settings.DEBUG))
        except Exception as kakao_err:
            print(f"[카카오 견적서 알림 오류] {kakao_err}")

        if is_ajax:
            return JsonResponse({"success": True, "message": f"견적서를 {', '.join(recipients)} 로 발송했습니다."})
        messages.success(request, f"견적서를 {', '.join(recipients)} 로 발송했습니다.")

    except Exception as e:
        if is_ajax:
            return JsonResponse({"success": False, "message": f"메일 발송 실패: {e}"}, status=500)
        messages.error(request, f"메일 발송 실패: {e}")

    return redirect('release_list')

@user_passes_test(is_admin)
def toggle_payment_status(request, release_id):
    release = get_object_or_404(MaterialRelease, id=release_id)

    if release.payment_status == 'paid':
        release.payment_status = 'unpaid'
        messages.warning(request, f"{release.institution.name} {release.order_month} 미수금 처리되었습니다.")
    else:
        release.payment_status = 'paid'
        messages.success(request, f"{release.institution.name} {release.order_month} 수금 완료되었습니다.")

    release.save()
    from .utils import redirect_with_filters
    return redirect_with_filters(request, "release_list")

@login_required
@user_passes_test(is_admin)
@require_POST
def release_material_item(request, item_id):
    """관리자만 품목 출고/취소 토글"""
    item = get_object_or_404(MaterialReleaseItem, id=item_id)
    material = item.material
    from robot_LvUP.views import find_levelup_release, sync_levelup_shipment_status

    if item.status == 'released':
        # ✅ 출고 취소: 센터수령만 재고 복구 + 상태 되돌림
        if item.release_method == '센터수령' and material and hasattr(material, 'stock'):
            old_stock = material.stock
            material.stock = F('stock') + (item.quantity or 0)
            material.save(update_fields=['stock'])
            material.refresh_from_db(fields=["stock"])  # 최신값 반영

            # ✅ 기록 남기기
            from .utils import log_material_history
            log_material_history(
                material=material,
                user=request.user,
                change_type="stock_restore",
                old_value=old_stock,
                new_value=material.stock,
                note=f"출고 취소 (출고품목ID {item.id}, 수량 {item.quantity})"
            )

        item.status = 'pending'
        item.released_at = None
        item.released_quantity = 0
        item.save(update_fields=['status', 'released_at', 'released_quantity'])

        # ✅ 출고 취소 시 연동된 주문(MaterialOrderItem) 자동 삭제
        MaterialOrderItem.objects.filter(notes__contains=f"[출고연동:{item.id}]").delete()
        matched_release = find_levelup_release(item.release.institution, item.release.order_month)
        if matched_release and matched_release.id == item.release.id:
            sync_levelup_shipment_status(item.release)
        messages.success(request, "출고 완료를 취소했습니다.")

    else:
        # ✅ 출고 처리
        if item.release_method == '센터수령' and material and hasattr(material, 'stock'):
            if (material.stock or 0) < (item.quantity or 0):
                messages.error(request, f"재고가 부족합니다. (현재 재고: {material.stock})")
                return redirect('release_list')

            old_stock = material.stock
            material.stock = F('stock') - (item.quantity or 0)
            material.save(update_fields=['stock'])
            material.refresh_from_db(fields=["stock"])  # 최신값 반영

            # ✅ 기록 남기기
            from .utils import log_material_history
            log_material_history(
                material=material,
                user=request.user,
                change_type="stock_decrease",
                old_value=old_stock,
                new_value=material.stock,
                note=f"출고 처리 (출고품목ID {item.id}, 수량 {item.quantity})"
            )

        item.status = 'released'
        item.released_at = timezone.now()
        item.released_quantity = item.quantity
        item.save(update_fields=['status', 'released_at', 'released_quantity'])
        
        # ✅ 출고 시 주문(MaterialOrder) 내역 자동 생성
        order_date = timezone.now().date()
        teacher = item.release.teacher
        order, created = MaterialOrder.objects.get_or_create(
            teacher=teacher,
            ordered_date=order_date,
            receive_type='order',
            defaults={
                'notes': f"{item.release.institution.name} 출고에 의한 자동생성"
            }
        )
        MaterialOrderItem.objects.create(
            order=order,
            vendor=item.vendor,
            material=item.material,
            quantity=item.quantity,
            status='waiting',
            notes=f"[출고연동:{item.id}]" + (" [택배]" if item.release_method == '택배' else " [센터수령]")
        )

        matched_release = find_levelup_release(item.release.institution, item.release.order_month)
        if matched_release and matched_release.id == item.release.id:
            sync_levelup_shipment_status(item.release)
        messages.success(request, "출고를 완료했습니다.")

    from .utils import redirect_with_filters
    return redirect_with_filters(request, "release_list")




@login_required
@user_passes_test(is_admin)
def unrelease_material_item(request, item_id):
    """관리자만 출고 취소(되돌리기)"""
    item = get_object_or_404(MaterialReleaseItem, id=item_id)
    from robot_LvUP.views import find_levelup_release, sync_levelup_shipment_status
    if request.method != 'POST':
        messages.error(request, '잘못된 요청입니다.')
        return redirect('release_list')

    if item.status != 'released':
        messages.info(request, '출고완료 상태가 아닙니다.')
        return redirect('release_list')

    # ✅ 되돌리기
    item.status = 'pending'
    item.released_at = None
    item.released_quantity = 0
    item.save()

    # ✅ 출고 취소 시 연동된 주문(MaterialOrderItem) 자동 삭제
    MaterialOrderItem.objects.filter(notes__contains=f"[출고연동:{item.id}]").delete()

    matched_release = find_levelup_release(item.release.institution, item.release.order_month)
    if matched_release and matched_release.id == item.release.id:
        sync_levelup_shipment_status(item.release)

    # ✅ 품목별 출고방법 체크
    if item.release_method == '센터수령':
        material = item.material
        if hasattr(material, 'stock'):
            material.stock = F('stock') + item.quantity
            material.save(update_fields=['stock'])

    messages.success(request, '출고가 취소되었습니다.')
    return redirect('release_list')

@login_required
def edit_release_item(request, item_id):
    item = get_object_or_404(MaterialReleaseItem, id=item_id)
    if item.release.source_type == 'levelup_auto':
        messages.warning(request, '단계업관리에서 자동 생성된 출고는 교구출고에서 직접 수정할 수 없습니다.')
        return redirect_with_filters(request, 'release_list')
    vendors = Vendor.objects.all()
    # 🔽 거래처별 정렬
    materials = Material.objects.all().order_by("vendor", "vendor_order", "name")

    # 🔹 현재 필터 값 (release_list에서 넘어온 GET 파라미터)
    selected_teacher_id = request.GET.get("teacher")
    selected_institution_id = request.GET.get("institution")
    selected_order_month = request.GET.get("order_month")

    if request.method == 'POST':
        vendor_id = request.POST.get('vendor')
        material_id = request.POST.get('material')
        unit_price = request.POST.get('unit_price')
        quantity = request.POST.get('quantity')
        release_method = request.POST.get('release_method')

        # ✅ 주문년도 / 월 (출고 모델용)
        order_year = request.POST.get('order_year')
        order_month = request.POST.get('order_month')

        if vendor_id:
            item.vendor = Vendor.objects.get(id=vendor_id)
        if material_id:
            item.material = Material.objects.get(id=material_id)

        item.unit_price = int(unit_price or 0)
        item.quantity = int(quantity or 0)
        item.release_method = release_method or item.release_method

        if not item.group_name:
            vendor_type_name = item.vendor.vendor_type.name if item.vendor and item.vendor.vendor_type else ""
            item.group_name = extract_group_name(item.material.name, vendor_type_name)
            item.save(update_fields=["group_name"])

        # ✅ MaterialRelease에 반영
        if order_year and order_month:
            item.release.order_month = f"{order_year}-{int(order_month):02d}"
            item.release.save()

        try:
            with transaction.atomic():
                item.save()
            messages.success(request, '출고 품목이 수정되었습니다.')

            # ✅ 필터 유지 redirect
            from materials.utils import redirect_with_filters
            return redirect_with_filters(request, "release_list")
        except Exception as e:
            messages.error(request, f"수정 중 오류가 발생했습니다: {e}")

    return render(request, 'release/edit_release_item.html', {
        'item': item,
        'vendors': vendors,
        'materials': materials,
        # 🔹 필터 값 템플릿에 넘김 (취소 버튼 등에서 사용)
        'selected_teacher_id': selected_teacher_id,
        'selected_institution_id': selected_institution_id,
        'selected_order_month': selected_order_month,
    })


@login_required
def delete_release_item(request, item_id):
    item = get_object_or_404(MaterialReleaseItem, id=item_id)
    if item.release.source_type == 'levelup_auto':
        messages.warning(request, '단계업관리에서 자동 생성된 출고는 교구출고에서 직접 삭제할 수 없습니다.')
        return redirect_with_filters(request, 'release_list')
    release = item.release  # 삭제 전 보관

    # 출고완료 & 센터수령이면 재고 복원
    if item.status == 'released' and item.release_method == '센터수령':
        material = item.material
        if hasattr(material, 'stock'):
            material.stock = F('stock') + item.quantity
            material.save(update_fields=['stock'])

    item.delete()

    # 남은 품목이 없으면 헤더도 삭제 (→ 해당 주문년월 객체도 사라짐)
    if not release.items.exists():
        release.delete()

    messages.success(request, '출고 품목이 삭제되었습니다.')
    from .utils import redirect_with_filters
    return redirect_with_filters(request, "release_list")

@login_required
@user_passes_test(is_admin)
@require_POST
def delete_release_group(request, institution_id, order_month):
    # 같은 기관 + 같은 주문월에 해당하는 모든 release
    releases = (MaterialRelease.objects
                .filter(institution_id=institution_id, order_month=order_month)
                .prefetch_related('items__material'))

    if releases.filter(source_type='levelup_auto').exists():
        messages.warning(request, '단계업관리에서 자동 생성된 출고 그룹은 교구출고에서 직접 삭제할 수 없습니다.')
        return redirect('release_list')

    if not releases.exists():
        messages.info(request, '삭제할 내역이 없습니다.')
        return redirect('release_list')

    with transaction.atomic():
        # 모든 아이템 순회하며 재고 복원 필요 시 복원
        items = MaterialReleaseItem.objects.filter(release__in=releases).select_related('material')
        for it in items:
            if it.status == 'released' and it.release_method == '센터수령':
                material = it.material
                if hasattr(material, 'stock'):
                    material.stock = F('stock') + it.quantity
                    material.save(update_fields=['stock'])

        # 아이템 전부 삭제
        MaterialReleaseItem.objects.filter(release__in=releases).delete()
        # 헤더(주문년월 단위) 전부 삭제
        releases.delete()

    messages.success(request, f'{order_month} 주문 내역이 모두 삭제되었습니다.')
     # ✅ 필터 유지해서 redirect
    query_string = request.GET.urlencode()
    url = reverse("release_list")
    if query_string:
        url += f"?{query_string}"
    return redirect(url)

def material_bulk_delete(request):
    if request.method == 'POST':
        material_ids = request.POST.getlist('material_ids') # 'material_ids'는 HTML input의 name 속성
        Material.objects.filter(pk__in=material_ids).delete()
        return redirect('material_list') # 삭제 후 목록 페이지로 리다이렉트
    return redirect('material_list') # GET 요청 시 목록 페이지로 리다이렉트 (실제로는 POST로만 동작)

from datetime import datetime
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import MaterialOrder, MaterialOrderItem
from django.utils import timezone

# ✅ 관리자만 접근 가능
def is_admin(user):
    return user.is_staff or user.is_superuser


@login_required
@user_passes_test(is_admin)
def toggle_payment_group(request, grouper):
    """주문 그룹 단위로 입금 상태 토글"""
    try:
        date_obj = datetime.strptime(grouper, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "잘못된 날짜 형식입니다.")
        return redirect("order_list")

    orders = MaterialOrder.objects.filter(ordered_date=date_obj).prefetch_related("items")

    if not orders.exists():
        messages.error(request, "해당 주문 그룹을 찾을 수 없습니다.")
        return redirect("order_list")

    # 첫 번째 아이템의 상태를 기준으로 토글
    first_item = orders.first().items.first()
    if first_item and first_item.paid_date:
        # ✅ 이미 입금완료 → 미입금으로 취소
        for order in orders:
            for item in order.items.all():
                item.paid_date = None
                item.save()
        messages.success(request, f"{grouper} 주문건 입금완료가 취소되었습니다.")
    else:
        # ✅ 미입금 상태 → 입금완료 처리
        now = timezone.now().date()
        for order in orders:
            for item in order.items.all():
                item.paid_date = now
                item.save()
        messages.success(request, f"{grouper} 주문건이 입금완료 처리되었습니다.")

    return redirect("order_list")


from django.db.models import F
from django.utils import timezone
from datetime import datetime
from .utils import log_material_history   # ✅ 추가

@login_required
@user_passes_test(is_admin)
def receive_group(request, grouper):
    """주문 그룹 단위로 입고 상태 토글"""
    try:
        date_obj = datetime.strptime(grouper, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "잘못된 날짜 형식입니다.")
        return redirect("order_list")

    orders = MaterialOrder.objects.filter(ordered_date=date_obj).prefetch_related("items__material")

    if not orders.exists():
        messages.error(request, "해당 주문 그룹을 찾을 수 없습니다.")
        return redirect("order_list")

    first_item = orders.first().items.first()
    if first_item and first_item.status == "received":
        # ✅ 입고완료 → 미입고로 취소 (재고 되돌림)
        for order in orders:
            for item in order.items.all():
                if item.status == "received" and item.material:
                    material = item.material
                    old_stock = material.stock
                    material.stock = F("stock") - item.quantity
                    material.save(update_fields=["stock"])
                    material.refresh_from_db(fields=["stock"])

                    # 기록 남기기
                    log_material_history(
                        material=material,
                        user=request.user,
                        change_type="stock_restore",
                        old_value=old_stock,
                        new_value=material.stock,
                        note=f"입고 취소 (주문ID {order.id}, 품목ID {item.id}, 수량 {item.quantity})"
                    )

                item.status = "waiting"
                item.received_date = None
                item.save()
        messages.success(request, f"{grouper} 주문건의 입고완료가 취소되었습니다.")

    else:
        # ✅ 미입고 → 입고완료 처리 (재고 증가)
        now = timezone.now().date()
        for order in orders:
            for item in order.items.all():
                if item.status != "received" and item.material:
                    material = item.material
                    old_stock = material.stock
                    material.stock = F("stock") + item.quantity
                    material.save(update_fields=["stock"])
                    material.refresh_from_db(fields=["stock"])

                    # 기록 남기기
                    log_material_history(
                        material=material,
                        user=request.user,
                        change_type="stock_increase",
                        old_value=old_stock,
                        new_value=material.stock,
                        note=f"입고 처리 (주문ID {order.id}, 품목ID {item.id}, 수량 {item.quantity})"
                    )

                item.status = "received"
                item.received_date = now
                item.save()
        messages.success(request, f"{grouper} 주문건이 입고완료 처리되었습니다.")

    return redirect("order_list")


@login_required
@user_passes_test(is_admin)
@transaction.atomic
def delete_group_orders(request, grouper):
    """주문 그룹 단위 전체 삭제 (주문일자별 삭제)"""
    try:
        date_obj = datetime.strptime(grouper, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "잘못된 날짜 형식입니다.")
        return redirect("order_list")

    orders = MaterialOrder.objects.filter(ordered_date=date_obj).prefetch_related("items")

    if not orders.exists():
        messages.error(request, "해당 주문 그룹을 찾을 수 없습니다.")
        return redirect("order_list")

    for order in orders:
        order.items.all().delete()
        order.delete()

    messages.success(request, f"{grouper} 주문 그룹이 전체 삭제되었습니다.")
    return redirect("order_list")

@login_required
def tax_invoice_issue(request, institution_id, order_month):
    MaterialRelease.objects.filter(
        institution_id=institution_id,
        order_month=order_month
    ).update(tax_invoice_sent=True)

    messages.success(request, f"{order_month} 세금계산서 발행 처리 완료!")

    # 🔹 공통 유틸 함수로 redirect
    from .utils import redirect_with_filters
    return redirect_with_filters(request, "release_list")



def _vendor_group_qs(date_str, vendor_type_id, vendor_id):
    return (MaterialOrderItem.objects
            .select_related('order', 'vendor', 'vendor__vendor_type', 'material')
            .filter(
                order__ordered_date=date_str,   # ✅ 여기 수정
                vendor__vendor_type_id=vendor_type_id,
                vendor_id=vendor_id,
            ))

@login_required
@require_POST
def toggle_payment_vendor_group(request, date_str, vendor_type_id, vendor_id):
    qs = _vendor_group_qs(date_str, vendor_type_id, vendor_id)

    already_paid = qs.filter(paid_date__isnull=False).exists()
    if already_paid:
        qs.update(paid_date=None)
        messages.success(request, "해당 거래처 그룹의 입금완료를 취소했습니다.")
    else:
        qs.update(paid_date=timezone.now().date())
        messages.success(request, "해당 거래처 그룹을 입금완료 처리했습니다.")
    return redirect('order_list')


def _date_group_qs(date_str):
    date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    return MaterialOrderItem.objects.filter(
        order__ordered_date=date_obj,
        order__receive_type='order'
    )

@login_required
@require_POST
def toggle_payment_date_group(request, date_str):
    qs = _date_group_qs(date_str)
    already_paid = qs.filter(paid_date__isnull=False).exists()
    if already_paid:
        qs.update(paid_date=None)
        messages.success(request, f"{date_str} 전체 주문의 입금완료를 취소했습니다.")
    else:
        qs.update(paid_date=timezone.now().date())
        messages.success(request, f"{date_str} 전체 주문을 입금완료 처리했습니다.")
    return redirect(request.META.get('HTTP_REFERER', 'order_list'))

@login_required
@require_POST
def toggle_receive_date_group(request, date_str):
    qs = _date_group_qs(date_str).select_related("material")
    already_received = qs.filter(status='received').exists()

    with transaction.atomic():
        if already_received:
            for item in qs:
                if item.status == "received":
                    material = item.material
                    if material and hasattr(material, 'stock'):
                        Material.objects.filter(id=material.id).update(
                            stock=F("stock") - (item.quantity or 0)
                        )
            qs.update(status="waiting", received_date=None)
            messages.success(request, f"{date_str} 전체 주문의 입고완료를 취소했습니다.")
        else:
            for item in qs:
                if item.status != "received":
                    material = item.material
                    if material and hasattr(material, 'stock'):
                        Material.objects.filter(id=material.id).update(
                            stock=F("stock") + (item.quantity or 0)
                        )
            qs.update(status="received", received_date=timezone.now().date())
            messages.success(request, f"{date_str} 전체 주문을 입고완료 처리했습니다.")

    return redirect(request.META.get('HTTP_REFERER', 'order_list'))


@login_required
@require_POST
def toggle_receive_vendor_group(request, date_str, vendor_type_id, vendor_id):
    qs = _vendor_group_qs(date_str, vendor_type_id, vendor_id).select_related("material")

    already_received = qs.filter(status='received').exists()

    with transaction.atomic():
        if already_received:
            # ▶ 입고 완료 → 미입고로 취소 (재고 되돌림)
            for item in qs:
                if item.status == "received":
                    material = item.material
                    if material and hasattr(material, 'stock'):
                        old_stock = material.stock
                        Material.objects.filter(id=material.id).update(
                            stock=F("stock") - (item.quantity or 0)
                        )
                        material.refresh_from_db(fields=["stock"])

                        log_material_history(
                            material=material,
                            user=request.user,
                            change_type="입고 취소",
                            old_value=old_stock,
                            new_value=material.stock,
                            note=f"거래처 그룹 입고취소 (item_id={item.id}, 수량 {item.quantity})"
                        )

                    item.status = "waiting"
                    item.received_date = None
                    item.save(update_fields=["status", "received_date"])
            messages.success(request, "해당 거래처 그룹의 입고완료가 취소되었습니다.")

        else:
            # ▶ 미입고 → 입고 완료 처리 (재고 증가)
            now = timezone.now().date()
            for item in qs:
                if item.status != "received":
                    material = item.material
                    if material and hasattr(material, 'stock'):
                        old_stock = material.stock
                        Material.objects.filter(id=material.id).update(
                            stock=F("stock") + (item.quantity or 0)
                        )
                        material.refresh_from_db(fields=["stock"])

                        log_material_history(
                            material=material,
                            user=request.user,
                            change_type="입고 완료",
                            old_value=old_stock,
                            new_value=material.stock,
                            note=f"거래처 그룹 입고완료 (item_id={item.id}, 수량 {item.quantity})"
                        )

                item.status = "received"
                item.received_date = now
                item.save(update_fields=["status", "received_date"])
            messages.success(request, "해당 거래처 그룹을 입고완료 처리했습니다.")

    return redirect("order_list")

@login_required
@user_passes_test(is_admin)
@require_POST
def delete_vendor_group_orders(request, date_str, vendor_type_id, vendor_id):
    qs = _vendor_group_qs(date_str, vendor_type_id, vendor_id)
    affected_orders = {item.order for item in qs}
    count = qs.count()

    for item in qs:
        order = item.order
        if order.receive_type == "return":
            # ✅ 반납입고 → 출고/재고 되돌리기
            material = item.material
            qty = item.quantity

            if material:
                old_stock = material.stock

                # 출고항목 복원
                release_item = MaterialReleaseItem.objects.filter(
                    material=material,
                    vendor=item.vendor
                ).order_by("-released_at").first()
                if release_item:
                    release_item.quantity += qty
                    release_item.save()

                # 재고 되돌리기
                material.stock = (material.stock or 0) - qty
                if material.stock < 0:
                    material.stock = 0
                material.save()

                log_material_history(
                    material=material,
                    user=request.user,
                    change_type="stock_restore",
                    old_value=old_stock,
                    new_value=material.stock,
                    note=f"반납입고 처리 (item_id={item.id}, 수량 {qty})"
                )

        # ✅ 아이템 삭제
        item.delete()

    # ✅ 남은 아이템 없는 주문은 함께 삭제
    for order in affected_orders:
        if not order.items.exists():
            order.delete()

    messages.success(
        request,
        f"{count}건을 삭제했습니다. (반납입고는 출고/재고 원복 완료)"
    )
    return redirect("order_list")

        
@csrf_exempt
@login_required
@require_POST
def material_reorder(request):
    """거래처별 교구재 순서 저장 (드래그앤드롭)"""
    try:
        data = json.loads(request.body.decode("utf-8"))  # [{id:1, vendor:3, position:1}, ...]
        for item in data:
            Material.objects.filter(id=item["id"], vendor_id=item["vendor"]).update(
                vendor_order=item["position"]
            )
        return JsonResponse({"status": "success"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)
    

@transaction.atomic
def return_release_item(request, item_id):
    """출고 항목 반납 처리 → 출고수량 감소, 재고 증가, 입고 생성"""
    if request.method != "POST":
        messages.error(request, "잘못된 요청입니다.")
        return redirect("release_list")

    item = get_object_or_404(MaterialReleaseItem, pk=item_id)

    try:
        return_qty = int(request.POST.get("return_qty") or 0)
    except ValueError:
        return_qty = 0

    if return_qty <= 0:
        messages.error(request, "반납 수량은 1 이상이어야 합니다.")
        return redirect("release_list")

    if return_qty > item.quantity:
        messages.error(request, f"반납 수량({return_qty})이 출고 수량({item.quantity})보다 클 수 없습니다.")
        return redirect("release_list")

    # 1) 출고 수량 줄이기
    item.quantity -= return_qty
    item.save()

    # 2) 재고 증가
    material = item.material
    material.stock = (material.stock or 0) + return_qty
    material.save()

    # ✅ 강사명 가져오기
    teacher_name = ""
    if hasattr(item.release, "teacher") and item.release.teacher:
        teacher = item.release.teacher
        teacher_name = teacher.first_name or teacher.username

    # 3) 교구재 입고 기록 생성
    order = MaterialOrder.objects.create(
        teacher=item.release.teacher,            # 반납 처리자
        receive_type="return",               # 반납입고
        ordered_date=timezone.now().date(),
        expected_date=timezone.now().date(),
        notes=f"[자동반납] {teacher_name} / 출고항목 #{item.id} 반납 {return_qty}개"
    )

    MaterialOrderItem.objects.create(
        order=order,
        material=material,
        vendor=item.vendor,
        quantity=return_qty,
    )

    messages.success(request, f"{material.name} {return_qty}개 반납 처리 및 입고 기록 생성 완료")
    return redirect("release_list")

@login_required
def return_list(request):
    """강사/관리자용 반납 내역"""
    orders_query = MaterialOrder.objects.filter(receive_type="return").prefetch_related("items__vendor__vendor_type", "items__material")
    
    if not request.user.is_staff:
        orders_query = orders_query.filter(teacher=request.user)

    orders = orders_query.order_by("-ordered_date").distinct()

    date_map = {}
    date_order = []

    for order in orders:
        date_str = order.ordered_date.strftime("%Y-%m-%d")

        if date_str not in date_map:
            date_map[date_str] = {
                "date": order.ordered_date,
                "date_str": date_str,
                "orders": [],
                "daily_total_qty": 0,
                "daily_total_sum": 0,
            }
            date_order.append(date_str)

        vendors_map = {}
        order_has_items = False
        
        for item in order.items.all():
            if item.receive_type != 'return':
                continue
                
            order_has_items = True
            vtype = item.vendor.vendor_type.name if (item.vendor and item.vendor.vendor_type) else "미지정"
            vname = item.vendor.name if item.vendor else "미지정"
            vtype_id = item.vendor.vendor_type_id if (item.vendor and item.vendor.vendor_type_id) else 0
            vid = item.vendor_id if item.vendor_id else 0

            vkey = f"{vtype} - {vname}"
            bucket = vendors_map.get(vkey)
            if not bucket:
                bucket = {
                    "vendor_type": vtype,
                    "vendor_name": vname,
                    "vendor_type_id": vtype_id,
                    "vendor_id": vid,
                    "vkey": vkey,
                    "items": [],
                    "unit_sum": 0,
                    "total_sum": 0,
                    "qty_sum": 0,
                }
                vendors_map[vkey] = bucket

            unit = (item.material.supply_price or 0)
            qty = (item.quantity or 0)
            bucket["items"].append(item)
            bucket["unit_sum"] += unit
            bucket["total_sum"] += unit * qty
            bucket["qty_sum"] += qty
            
            # 일자별 누적
            date_map[date_str]["daily_total_qty"] += qty
            date_map[date_str]["daily_total_sum"] += unit * qty

        if order_has_items:
            order_block = {
                "order": order,
                "vendors": sorted(
                    vendors_map.values(),
                    key=lambda b: (b["vendor_type"], b["vendor_name"])
                ),
            }
            date_map[date_str]["orders"].append(order_block)

    date_blocks = [date_map[ds] for ds in date_order]

    return render(request, "release/return_list.html", {
        "date_blocks": date_blocks
    })


@login_required
def release_item_students(request, item_id):
    """교구출고 품목의 연결된 단계업 학생 목록 반환 (JSON)"""
    from robot_LvUP.models import RobotLevelUp
    item = get_object_or_404(MaterialReleaseItem, id=item_id)
    release = item.release

    students = list(
        RobotLevelUp.objects.filter(
            institution=release.institution,
            year_month=release.order_month,
            material=item.material,
        ).select_related("student__division").order_by("section", "grade", "class_no", "student_no").values(
            "id", "student_name", "section", "grade", "class_no", "student_no",
            "shipped_done", "delivery_done",
            "student__division__division",
        )
    )

    result = []
    for s in students:
        section = s["student__division__division"] or s["section"] or ""
        result.append({
            "id": s["id"],
            "name": s["student_name"],
            "section": section,
            "grade": s["grade"] or "",
            "class_no": s["class_no"] or "",
            "student_no": s["student_no"] or "",
            "shipped": s["shipped_done"],
            "delivered": s["delivery_done"],
        })

    return JsonResponse({
        "material": item.material.name,
        "total": item.quantity,
        "released_quantity": item.released_quantity,
        "students": result,
    })
